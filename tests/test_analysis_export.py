from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook

from forecast.analysis_export import build_comparison_audit_workbook
from forecast.sales_comparison import calculate_sales_effect_rows, sales_effect_totals


class FakeGoldenWorkbook:
    def __init__(self, path):
        self.path = str(path)
        self.formulas = {"E10": "=1+1", "E20": "=E10*2"}

    def value(self, cell):
        return 200.0 if "comparison" in self.path else 100.0


def test_build_comparison_audit_workbook(monkeypatch, tmp_path):
    monkeypatch.setattr("forecast.analysis_export.GoldenWorkbook", FakeGoldenWorkbook)
    mapping_path = tmp_path / "mapping.json"
    mapping_path.write_text(json.dumps({
        "comparison": {
            "pnl_rows": {"operating_profit": 10},
            "pnl_labels": {"operating_profit": "영업이익"},
            "products": {},
            "sales_groups": {"SW": {"label": "SW", "quantity_row": 10, "amount_row": 20, "cogs_row": 30}},
            "production_rows": {"SW400": 40},
            "production_labels": {"SW400": "SW400 생산량"},
            "mcm_rows": {"SW400": 50},
            "mcm_labels": {"SW400": "MCM SW400"},
            "cost_rows": {"raw_material": {"add": [60], "subtract": [61]}},
            "cost_labels": {"raw_material": "원재료비"},
            "effect_rows": {"revenue": 70},
            "effect_labels": {"revenue": "매출액"},
        },
        "analysis_adapter": {
            "material": {
                "jpy_fx_row": 9,
                "front_process": {
                    "nonwoven_quantity_row": 205,
                    "nonwoven_amount_row": 206,
                    "nonwoven_unit_row": 207,
                    "other_quantity_row": 208,
                    "other_amount_row": 209,
                    "other_unit_row": 210,
                    "total_amount_row": 211
                },
                "back_process": {"source_start_row": 684, "source_end_row": 699}
            },
            "account_discovery": {"manufacturing_start_marker": "제조경비"},
            "manufacturing": {
                "front_ratio_rows": {"labor": 345, "outsourcing": 346, "other_variable": 347}
            }
        }
    }), encoding="utf-8")

    result = {
        "baseline": {"id": "base", "name": "기준", "model_type": "계획", "version": "V1"},
        "comparison": {"id": "comp", "name": "비교", "model_type": "실적", "version": "V1"},
        "period": {"label": "1월", "months": [1]},
        "pnl": [{"code": "operating_profit", "item": "영업이익", "baseline": 1000, "comparison": 1200, "delta": 200}],
        "effects": [{"code": "revenue", "factor": "매출액", "baseline": 1000, "comparison": 1200, "delta": 200, "profit_effect": 200}],
        "operating_profit_delta": 200,
        "effects_total": 200,
        "residual": 0,
        "reconciled": True,
        "cost_summary": [
            {"code": "raw_material", "item": "원재료비", "baseline": 300, "comparison": 350, "delta": 50},
            {"code": "selling_expense", "item": "판매비", "baseline": 100, "comparison": 90, "delta": -10},
        ],
        "mcm": [{"code": "SW400", "item": "MCM SW400", "baseline": 10, "comparison": 20, "delta": 10}],
        "production": [{"code": "SW400", "item": "SW400 생산량", "baseline": 100, "comparison": 110, "delta": 10}],
        "material_analysis": {
            "product_groups": [{
                "product_group": "SW", "baseline_unit_cost": 10, "comparison_unit_cost": 12,
                "unit_cost_delta": 2, "nonwoven_price_ex_fx": -50,
                "nonwoven_jpy": -30, "materials_ex_nonwoven": -120,
                "total": -200, "calculation_status": "완료",
            }],
        },
        "manufacturing_accounts": [{
            "row": 297, "account": "수도광열비", "classification": "variable",
            "allocation_ratio_row": 347, "baseline_front_ratios": [0.54],
            "baseline_amount": 100, "comparison_amount": 90, "delta": -10,
            "activity_effect": 4, "unit_effect": 6, "fixed_effect": 0,
            "occurrence_effect": 10, "inventory_realization_rate": 1.1,
            "final_profit_effect": 11, "calculation_status": "완료",
        }],
        "sga_accounts": [{
            "row": 1168, "section": "판매비", "account": "운반비",
            "classification": "transport", "baseline_amount": 10,
            "comparison_amount": 20, "delta": 10, "profit_effect": 0,
            "bridge_position": "판매효과",
        }],
    }
    sales_input = [{
        "product_group": "SW",
        "baseline_quantity": 100,
        "baseline_amount": 1_000_000,
        "baseline_gross_margin_rate": 0.3,
        "comparison_quantity": 110,
        "comparison_amount": 1_210_000,
        "comparison_gross_margin_rate": 0.32,
    }]
    sales_rows = calculate_sales_effect_rows(sales_input, 1400.0, 1450.0)
    totals = sales_effect_totals(sales_rows)

    payload = build_comparison_audit_workbook(
        result=result,
        sales_rows=sales_rows,
        sales_totals=totals,
        baseline_fx=1400.0,
        comparison_fx=1450.0,
        baseline_path=tmp_path / "baseline.xlsx",
        comparison_path=tmp_path / "comparison.xlsx",
        mapping_path=mapping_path,
    )

    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert "원천셀_추적" in workbook.sheetnames
    assert workbook["판매효과_검증"]["D5"].value.startswith("=")
    assert workbook["손익_정합성"]["G5"].value == "=E5-D5"
    assert workbook["원천셀_추적"].max_row > 4
    trace_cells = {
        cell.value
        for row in workbook["원천셀_추적"].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("E")
    }
    assert {"E9", "E205", "E211", "E684", "E699", "E289", "E319", "E345", "E347"} <= trace_cells
    assert workbook["README"]["B15"].value == "PASS"
    material_text = " ".join(
        str(cell.value or "") for row in workbook["원부재료_검증"].iter_rows() for cell in row
    )
    assert "MCM SW400" not in material_text
    assert workbook["원부재료_검증"]["I5"].value == "=SUM(E5:G5)"
    assert workbook["생산제조경비_검증"]["O5"].value == "=SUM(I5:K5)"
    assert workbook["생산제조경비_검증"]["D5"].value == 347
    assert workbook["판관비_검증"]["I5"].value == "판매효과"
