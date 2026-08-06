from __future__ import annotations

import json
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from forecast.comparison import GenericComparisonEngine
from forecast.storage import ModelRegistry
from forecast.workbook import extract_period_types, infer_workbook_year


ROOT = Path(__file__).resolve().parents[1]


def write_metadata_workbook(path: Path, year: int | None = 2027) -> bytes:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    data = workbook.create_sheet("Data")
    if year is not None:
        for month in range(1, 13):
            data.cell(2, month + 4).value = date(year, month, 1)
    statuses = ["실적"] * 6 + ["추정"] * 3 + ["계획"] * 3
    for month, status in enumerate(statuses, start=1):
        data.cell(3, month + 4).value = status
    workbook.save(path)
    return path.read_bytes()


class ModelMetadataTests(unittest.TestCase):
    def test_period_types_are_read_from_data_row_three(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "model.xlsx"
            write_metadata_workbook(path)

            period_types = extract_period_types(path)

        self.assertEqual(period_types["1"], "실적")
        self.assertEqual(period_types["6"], "실적")
        self.assertEqual(period_types["7"], "추정")
        self.assertEqual(period_types["10"], "계획")
        self.assertEqual(len(period_types), 12)

    def test_year_is_inferred_from_date_headers_and_uses_explicit_fallback(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "dated.xlsx"
            write_metadata_workbook(path, year=2031)
            self.assertEqual(infer_workbook_year(path, fallback_year=1999), 2031)

            fallback_path = Path(directory) / "no_year.xlsx"
            write_metadata_workbook(fallback_path, year=None)
            self.assertEqual(infer_workbook_year(fallback_path, fallback_year=1999), 1999)

    def test_registry_defaults_to_full_year_and_persists_period_types(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "model.xlsx"
            content = write_metadata_workbook(workbook_path, year=2032)
            registry = ModelRegistry(Path(directory) / "registry")
            meta = registry.add(
                content,
                name="테스트 모형",
                model_type="계획",
                created_date="2032-01-01",
                version="V1",
                confirmed=True,
                file_name="model.xlsx",
            )

            self.assertEqual(meta.year, 2032)
            self.assertEqual((meta.start_month, meta.end_month), (1, 12))
            self.assertEqual(meta.period_types["7"], "추정")
            self.assertIn("1~6월 실적 / 7~9월 추정 / 10~12월 계획", meta.period_composition)

            saved = json.loads(registry.index.read_text(encoding="utf-8"))[0]
            self.assertEqual(saved["start_month"], 1)
            self.assertEqual(saved["end_month"], 12)
            self.assertEqual(saved["period_types"]["12"], "계획")

    def test_legacy_models_json_without_period_types_remains_readable(self):
        with tempfile.TemporaryDirectory() as directory:
            registry = ModelRegistry(directory)
            registry.index.write_text(
                json.dumps([{
                    "id": "legacy",
                    "name": "기존 모형",
                    "model_type": "계획",
                    "year": 2026,
                    "start_month": 1,
                    "end_month": 12,
                    "created_date": "2026-01-01",
                    "version": "V1",
                    "confirmed": True,
                    "file_name": "legacy.xlsx",
                    "uploaded_at": "2026-01-01T00:00:00+09:00",
                }], ensure_ascii=False),
                encoding="utf-8",
            )

            models = registry.list()

        self.assertEqual(len(models), 1)
        self.assertEqual(models[0].period_types, {})
        self.assertEqual(models[0].period_composition, "미확인")

    def test_registered_models_keep_common_months_full_year(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "model.xlsx"
            content = write_metadata_workbook(workbook_path, year=2033)
            registry = ModelRegistry(Path(directory) / "registry")
            first = registry.add(content, name="기준", model_type="계획")
            second = registry.add(content, name="비교", model_type="추정")

            months = GenericComparisonEngine(ROOT / "config" / "model_mapping.json").common_months(first, second)

        self.assertEqual(months, tuple(range(1, 13)))


if __name__ == "__main__":
    unittest.main()
