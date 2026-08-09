from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from forecast.preflight import ExcelPreflightValidator, PreflightValidationError


ROOT = Path(__file__).resolve().parents[1]
MAPPING = json.loads((ROOT / "config/model_mapping.json").read_text(encoding="utf-8"))


def _workbook(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    for month, column in enumerate(range(5, 17), 1):
        sheet.cell(2, column, f"2026년 {month}월")
        sheet.cell(3, column, "계획")

    discovery = MAPPING["analysis_adapter"]["account_discovery"]
    sheet["C211"] = "생산출고 계"
    sheet["C699"] = "생산출고 계"
    sheet["B287"] = discovery["manufacturing_start_marker"]
    sheet["B321"] = discovery["manufacturing_stop_marker"]
    sheet["B1166"] = discovery["sga_start_marker"]
    sheet["B1244"] = discovery["sga_stop_marker"]
    sheet["B1306"] = "영업이익"

    validator = ExcelPreflightValidator(MAPPING)
    for spec in validator.numeric_rows:
        for row in spec.rows:
            for column in range(5, 17):
                sheet.cell(row, column, 0)
    book.save(path)


def _require(path: Path):
    return ExcelPreflightValidator(MAPPING).require(path, expected_year=2026)


def _mutate(path: Path, callback) -> None:
    book = load_workbook(path)
    callback(book)
    book.save(path)
    book.close()


def test_anchor_and_hierarchical_block_preflight_passes(tmp_path):
    path = tmp_path / "valid.xlsx"
    _workbook(path)

    report = _require(path)

    assert report.passed
    assert report.anchors == {
        "front_raw_material_total": 211,
        "back_raw_material_total": 699,
        "manufacturing_start": 287,
        "manufacturing_stop": 321,
        "sga_start": 1166,
        "sga_stop": 1244,
        "operating_profit": 1306,
    }


def test_irrelevant_stp_defined_name_merge_and_hidden_row_do_not_fail(tmp_path):
    path = tmp_path / "decorated.xlsx"
    _workbook(path)

    def decorate(book):
        stp = book.create_sheet("STP")
        stp["A1"] = "supplementary data"
        book.defined_names.add(DefinedName("STP_Check", attr_text="'STP'!$A$1"))
        data = book["Data"]
        data.merge_cells("R10:S10")
        data["R10"] = "not an analysis source"
        data.row_dimensions[1800].hidden = True

    _mutate(path, decorate)

    assert _require(path).passed


@pytest.mark.parametrize(
    ("name", "mutator", "expected_code"),
    [
        (
            "anchor label",
            lambda book: setattr(book["Data"]["C211"], "value", "합계 훼손"),
            "anchor_missing",
        ),
        (
            "anchor row move",
            lambda book: (
                setattr(book["Data"]["C211"], "value", None),
                setattr(book["Data"]["C212"], "value", "생산출고 계"),
            ),
            "anchor_missing",
        ),
        (
            "required row delete",
            lambda book: book["Data"].delete_rows(699, 1),
            "anchor_missing",
        ),
        (
            "month header",
            lambda book: setattr(book["Data"]["E2"], "value", "2026년 13월"),
            "month_header_invalid",
        ),
        (
            "numeric source",
            lambda book: setattr(book["Data"]["E9"], "value", "JPY 없음"),
            "source_cell_not_numeric",
        ),
        (
            "data sheet rename",
            lambda book: setattr(book["Data"], "title", "Data_Broken"),
            "data_sheet_missing",
        ),
    ],
)
def test_core_source_damage_is_blocked_before_calculation(
    tmp_path, name, mutator, expected_code
):
    path = tmp_path / f"broken-{name}.xlsx"
    _workbook(path)
    _mutate(path, mutator)

    with pytest.raises(PreflightValidationError) as captured:
        _require(path)

    assert any(issue.code == expected_code for issue in captured.value.report.issues)
    assert str(captured.value).strip()
