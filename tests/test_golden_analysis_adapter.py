from __future__ import annotations

import json

from openpyxl import Workbook, load_workbook

from forecast.analysis.configuration import AnalysisConfig
from forecast.analysis.golden_adapter import GoldenAnalysisAdapter
from forecast.storage import ModelMeta
from forecast.workbook import GoldenWorkbook


def _meta(identifier: str) -> ModelMeta:
    return ModelMeta(
        identifier, identifier, "계획", 2026, 1, 12,
        "2026-01-01", "V1", True, f"{identifier}.xlsx", "now",
    )


def _build_workbook(path, *, comparison: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["B287"] = "★제조경비 명세서_입력"
    sheet["C289"] = "노무비"
    sheet["D290"] = "급료"
    sheet["C296"] = "제조경비"
    for row, account in (
        (297, "수도광열비"),
        (298, "소모품비"),
        (304, "원자재운반비"),
        (305, "외주가공비"),
        (306, "신규 제조계정"),
    ):
        sheet.cell(row, 3, "변동비" if row in {297, 298, 304, 305} else "고정비")
        sheet.cell(row, 4, account)
    sheet["B321"] = "*제조원가 변동비/고정비 비율"

    sheet["B1166"] = "★판매관리비 관리 명세서"
    sheet["B1167"] = "판매비"
    for row, classification, account in (
        (1168, "변동비", "운반비"),
        (1170, "변동비", "브랜드사용료"),
        (1194, "변동비", "포장비"),
    ):
        sheet.cell(row, 2, classification)
        sheet.cell(row, 3, account)
    sheet["B1197"] = "일반관리비"
    for row, account in ((1198, "급여"), (1214, "운반비"), (1221, "신규 판관계정")):
        sheet.cell(row, 2, "고정비")
        sheet.cell(row, 3, account)
    sheet["B1244"] = "★손익계산서"

    column = 5
    def put(row: int, base: float, target: float | None = None):
        sheet.cell(row, column, target if comparison and target is not None else base)

    put(9, 10, 12)
    put(205, 1_000)
    put(206, 10_000, 12_000)
    put(208, 1_000)
    put(209, 0)
    put(211, 10_000, 12_000)
    put(425, 100)
    put(426, 1_000, 1_200)
    put(430, 0)
    put(431, 0)
    put(435, 0)
    put(436, 0)

    put(32, 100, 90)
    put(35, 1)
    put(38, 0)
    put(41, 1)
    put(1593, 100, 90)
    put(556, 60)
    put(557, 40)
    put(568, 10)
    put(569, 0)
    for row, value in ((897, 60), (904, 40), (580, 1), (583, 1)):
        put(row, value)
    for row in (956, 957):
        put(row, 1)
    put(273, 0.1)
    put(274, 0)
    put(275, 0)
    put(685, 300, 100)
    put(694, 0)
    put(697, 0)
    put(699, 300, 100)
    put(788, 0.5)
    put(789, 0.5)
    for row in (790, 791, 792):
        put(row, 0)
    for row, base, target in (
        (900, 50, 50), (907, 50, 50),
    ):
        put(row, base, target)

    # Other material groups are valid zero-activity rows in this fixture.
    for row in (
        44, 47, 50, 53, 56, 59, 62, 65, 68, 71, 74, 77,
        119, 122, 125, 558, 559, 560, 570, 571, 911, 918, 925,
        586, 589, 592, 958, 959, 960, 913, 914, 920, 921, 927,
        1632, 1720,
    ):
        put(row, 0)

    put(345, 0.5, 0.9)
    put(346, 0.2, 0.9)
    put(347, 0.4, 0.9)
    for row, base, target in (
        (290, 100, 110),
        (297, 1_000, 900),
        (298, 200, 180),
        (304, 100, 80),
        (305, 500, 450),
        (306, 50, 40),
    ):
        put(row, base, target)
    put(119, 100, 80)
    put(556, 60, 60)
    put(557, 40, 40)
    put(558, 0, 20)
    put(568, 10, 20)

    # Comparison realization rate = COGS / current-period manufacturing input = 1.
    put(1268, 200)
    put(1248, 300)
    put(1306, 25)
    put(1276, 100)
    put(1273, 0)
    put(1274, 0)
    put(1285, 0)
    put(1277, 25)
    put(1286, 25)
    put(1278, 25)
    put(1287, 0)
    put(1279, 25)
    put(1288, 0)

    for row, base, target in (
        (1168, 10, 20), (1170, 30, 25), (1194, 40, 35),
        (1198, 50, 45), (1214, 5, 6), (1221, 7, 8),
    ):
        put(row, base, target)
    workbook.save(path)


def _adapter() -> GoldenAnalysisAdapter:
    mapping = json.loads(open("config/model_mapping.json", encoding="utf-8").read())
    config = AnalysisConfig.load("config/analysis_v1.json")
    return GoldenAnalysisAdapter(mapping, config)


def test_adapter_calculates_material_three_part_identity_from_golden_cells(tmp_path):
    base_path = tmp_path / "base.xlsx"
    comparison_path = tmp_path / "comparison.xlsx"
    _build_workbook(base_path)
    _build_workbook(comparison_path, comparison=True)
    adapter = _adapter()
    base = adapter.build(GoldenWorkbook(base_path), _meta("base"), (1,))
    comparison = adapter.build(GoldenWorkbook(comparison_path), _meta("comparison"), (1,))
    result = adapter.material_analysis(base, comparison)
    sw = next(row for row in result["product_groups"] if row["product_group"] == "SW")

    assert sw["baseline_unit_cost"] == 14
    assert sw["comparison_unit_cost"] == 17
    assert sw["total"] == -270
    assert sw["nonwoven_price_ex_fx"] == 0
    assert sw["nonwoven_jpy"] == -180
    assert sw["materials_ex_nonwoven"] == -90
    assert sw["total"] == (
        sw["nonwoven_price_ex_fx"]
        + sw["nonwoven_jpy"]
        + sw["materials_ex_nonwoven"]
    )
    assert result["jpy_fx_unit"] == "KRW/JPY"
    assert "mcm" not in str(result).lower()
    assert "yield" not in str(result).lower()


def test_material_uses_direct_input_total_rows_211_and_699(tmp_path):
    base_path = tmp_path / "base.xlsx"
    comparison_path = tmp_path / "comparison.xlsx"
    _build_workbook(base_path)
    _build_workbook(comparison_path)
    workbook = load_workbook(comparison_path)
    workbook["Data"]["E211"] = 11_000
    workbook["Data"]["E699"] = 400
    workbook.save(comparison_path)
    adapter = _adapter()
    base = adapter.build(GoldenWorkbook(base_path), _meta("base"), (1,))
    comparison = adapter.build(GoldenWorkbook(comparison_path), _meta("comparison"), (1,))
    result = adapter.material_analysis(base, comparison)

    assert result["total"] < 0
    assert result["nonwoven_price_ex_fx"] == 0
    assert result["nonwoven_jpy"] == 0
    assert result["materials_ex_nonwoven"] == result["total"]


def test_adapter_calculates_all_manufacturing_accounts_with_baseline_ratios(tmp_path):
    base_path = tmp_path / "base.xlsx"
    comparison_path = tmp_path / "comparison.xlsx"
    _build_workbook(base_path)
    _build_workbook(comparison_path, comparison=True)
    adapter = _adapter()
    base = adapter.build(GoldenWorkbook(base_path), _meta("base"), (1,))
    comparison = adapter.build(GoldenWorkbook(comparison_path), _meta("comparison"), (1,))
    accounts, analysis = adapter.manufacturing_accounts(base, comparison)

    assert [row["account"] for row in accounts] == [
        "급료", "수도광열비", "소모품비", "원자재운반비", "외주가공비", "신규 제조계정",
    ]
    utilities = next(row for row in accounts if row["account"] == "수도광열비")
    assert utilities["classification"] == "variable"
    assert utilities["allocation_ratio_row"] == 347
    assert utilities["baseline_front_ratios"] == [0.4]
    assert utilities["activity_effect"] + utilities["unit_effect"] == 100
    assert utilities["occurrence_effect"] == (
        utilities["baseline_amount"] - utilities["comparison_amount"]
    )
    assert utilities["inventory_realization_rate"] == 1
    assert utilities["final_profit_effect"] == 100
    assert analysis["inventory_realization_rate"] == 1
    assert all("기타 제조경비" not in row["account"] for row in accounts)


def test_adapter_discovers_new_sga_rows_and_keeps_transport_sections_distinct(tmp_path):
    path = tmp_path / "model.xlsx"
    _build_workbook(path)
    adapter = _adapter()
    adapted = adapter.build(GoldenWorkbook(path), _meta("base"), (1,))
    rows = adapted.sga_source_rows

    assert any(row["account"] == "신규 판관계정" for row in rows)
    selling_transport = next(
        row for row in rows if row["account"] == "운반비" and row["section"] == "판매비"
    )
    general_transport = next(
        row for row in rows if row["account"] == "운반비" and row["section"] == "일반관리비"
    )
    assert selling_transport["row"] == 1168
    assert general_transport["row"] == 1214


def test_adapter_does_not_create_mixed_unit_transport_activity(tmp_path):
    path = tmp_path / "model.xlsx"
    _build_workbook(path)
    adapted = _adapter().build(
        GoldenWorkbook(path), _meta("base"), (1,), sales_fx=1500.0
    )

    assert any(row.unit_basis == "PCS" for row in adapted.scenario.products)
    assert any(row.unit_basis == "LENGTH" for row in adapted.scenario.products)
    assert adapted.scenario.activities[0].transport_activity == 0.0
    assert all(row.sales_fx == 1500.0 for row in adapted.scenario.products)
