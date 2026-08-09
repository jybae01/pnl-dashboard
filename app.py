import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
import logging
import re
import base64
from pathlib import Path

from forecast_dashboard.security import (
    clear_login_failures,
    load_credentials,
    lockout_remaining,
    register_failed_attempt,
    verify_access_code,
)
from forecast_dashboard.storage import (
    StorageError,
    delete_saved_data,
    load_saved_data,
    save_uploaded_data,
)
from forecast_dashboard.period_state import (
    apply_draft_period,
    get_applied_period,
    initialize_period_state,
)
from forecast_dashboard.workbooks import read_workbook, safe_extract

# 1. 페이지 기본 설정 (반드시 최상단에 위치)
st.set_page_config(page_title="손익계산서 조회", layout="wide")

APP_DIR = Path(__file__).resolve().parent
LOGGER = logging.getLogger(__name__)

def image_data_uri(filename):
    """실행 파일 옆의 브랜드 이미지를 화면에 안전하게 표시한다."""
    image_path = APP_DIR / filename
    if not image_path.exists():
        return ""
    encoded = base64.b64encode(image_path.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"

BLACK_LOGO_URI = image_data_uri("NanoH2O_Logo@3x_Black.png")
WHITE_LOGO_URI = image_data_uri("NanoH2O_Logo@3x_White.png")

try:
    ACCESS_CREDENTIALS = load_credentials(st.secrets)
except Exception as exc:
    st.error("접속 보안 설정이 완료되지 않았습니다. 관리자에게 문의해 주세요.")
    LOGGER.error("Invalid authentication configuration: %s", exc)
    st.stop()

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.role = None

if not st.session_state.authenticated:
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] {
        background:
            radial-gradient(circle at 15% 20%, rgba(255,107,44,.12), transparent 30%),
            radial-gradient(circle at 85% 82%, rgba(31,41,55,.08), transparent 32%),
            #FCFBFA;
    }
    [data-testid="stHeader"] { background: transparent; }
    .login-hero { text-align: center; margin: 15vh 0 22px; }
    .login-brand { width: 260px; margin: 0 auto 18px; padding: 17px 22px; border-radius: 14px;
        background: linear-gradient(135deg, #111111, #303030); box-shadow: 0 10px 22px rgba(17,17,17,.18); }
    .login-brand img { width: 100%; height: auto; display: block; }
    .login-hero h1 { margin: 0; color: #172033; font-size: 27px; letter-spacing: -.6px; }
    .login-hero p { margin: 8px 0 0; color: #64748B; font-size: 14px; }
    div[data-testid="stForm"] {
        max-width: 380px; margin: 0 auto; padding: 24px 24px 20px;
        border: 1px solid #E2E8F0; border-radius: 16px; background: rgba(255,255,255,.94);
        box-shadow: 0 18px 42px rgba(15,23,42,.10);
    }
    div[data-testid="stForm"] input {
        height: 44px; border-radius: 9px; border-color: #CBD5E1; font-size: 14px;
    }
    div[data-testid="stForm"] input:focus { border-color: #FF6B2C; box-shadow: 0 0 0 3px rgba(255,107,44,.14); }
    div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] > button,
    div[data-testid="stForm"] button[kind="secondaryFormSubmit"] {
        height: 42px; border: 0; border-radius: 9px; background: #FF6B2C; color: white;
        font-weight: 700; letter-spacing: -.1px;
    }
    div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] > button:hover,
    div[data-testid="stForm"] button[kind="secondaryFormSubmit"]:hover { background: #D9551E; }
    .login-help { text-align: center; color: #94A3B8; font-size: 12px; margin-top: 14px; }
    </style>
    <div class="login-hero">
        <div class="login-brand"><img src="__WHITE_LOGO_URI__" alt="NanoH2O"></div>
        <h1>손익 데이터 모니터링</h1>
        <p>접속 코드를 입력하여 대시보드를 확인하세요.</p>
    </div>
    """.replace("__WHITE_LOGO_URI__", WHITE_LOGO_URI), unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        with st.form("login_form"):
            entered_code = st.text_input("Access Code", type="password", label_visibility="collapsed", placeholder="접속 코드를 입력하세요")
            submitted = st.form_submit_button("접속", use_container_width=True)
            
            if submitted:
                remaining = lockout_remaining(st.session_state)
                if remaining:
                    st.error(f"로그인 시도가 잠시 제한되었습니다. {remaining}초 후 다시 시도해 주세요.")
                else:
                    role = verify_access_code(entered_code, ACCESS_CREDENTIALS)
                    if role:
                        clear_login_failures(st.session_state)
                        st.session_state.authenticated = True
                        st.session_state.role = role
                        st.rerun()
                    register_failed_attempt(st.session_state)
                    remaining = lockout_remaining(st.session_state)
                    if remaining:
                        st.error(f"로그인 시도가 잠시 제한되었습니다. {remaining}초 후 다시 시도해 주세요.")
                    else:
                        st.error("⚠️ 접속 코드가 일치하지 않습니다.")
        st.markdown("<div class='login-help'>권한이 필요한 경우 관리자에게 문의해 주세요.</div>", unsafe_allow_html=True)
    
    st.stop()
# ---------------------------------------------


# [HTML/CSS] 디자인 최적화 (라디오버튼 강제 중앙정렬 & YTD 표 자연스러운 간격 복구)
st.markdown("""
<style>
/* 조회 화면 공통 스타일 */
[data-testid="stAppViewContainer"] { background: #FCFBFA; }
[data-testid="stHeader"] { background: rgba(252,251,250,.9); }
.block-container { max-width: 1640px; padding-top: 2.1rem; padding-bottom: 3rem; }
.dashboard-brand {
    display: flex; align-items: center; gap: 20px; min-height: 62px;
    padding: 8px 0 14px; box-sizing: border-box;
    margin-bottom: 2px; border-bottom: 1px solid #E2E8F0; overflow: visible;
}
.dashboard-brand img { width: 188px; height: auto; display: block; }
.dashboard-brand-title { color: #172033; font-size: 27px; font-weight: 750; letter-spacing: -0.8px; line-height: 1.35; white-space: nowrap; }
h1 {
    color: #172033 !important; font-size: 30px !important; font-weight: 750 !important;
    letter-spacing: -0.8px !important; padding-bottom: 14px !important;
    border-bottom: 1px solid #E2E8F0;
}
h5 { color: #1E293B !important; font-weight: 700 !important; letter-spacing: -0.25px; }
[data-testid="stMetric"] {
    padding: 16px 18px; min-height: 132px; border: 1px solid #E2E8F0;
    border-radius: 14px; background: #FFFFFF; box-shadow: 0 6px 18px rgba(15,23,42,.045);
}
[data-testid="stMetricLabel"] { color: #64748B !important; font-size: 13px !important; font-weight: 650 !important; }
[data-testid="stMetricValue"] { color: #172033 !important; font-size: 24px !important; font-weight: 750 !important; }
[data-testid="stMetricDelta"] { font-size: 12px !important; white-space: normal !important; line-height: 1.45 !important; }
[data-testid="stSidebar"] { background: #FFFFFF; border-right: 1px solid #E2E8F0; }
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 { color: #242424; }
.stButton > button, [data-testid="stDownloadButton"] > button {
    border-radius: 8px; border-color: #CBD5E1; font-weight: 650; transition: all .15s ease;
}
.stButton > button:hover, [data-testid="stDownloadButton"] > button:hover {
    border-color: #FF6B2C; color: #D9551E; box-shadow: 0 3px 10px rgba(255,107,44,.14);
}
.stTabs [data-baseweb="tab-list"] { gap: 6px; border-bottom: 1px solid #E2E8F0; }
.stTabs [data-baseweb="tab"] {
    height: 38px; padding: 0 15px; border-radius: 8px 8px 0 0;
    color: #64748B; font-weight: 650;
}
.stTabs [aria-selected="true"] { color: #D9551E !important; background: #FFF0E9; }
hr { border-color: #E2E8F0 !important; margin: 1.4rem 0 !important; }

/* 💡 라디오 버튼 그룹 강제 중앙 정렬 */
div[role="radiogroup"] {
    display: flex !important;
    justify-content: center !important;
    margin-bottom: 5px !important;
}

/* 💡 드롭다운 사이즈 최소화 및 여백 제거 */
div[data-baseweb="select"] {
    font-size: 13px !important;
}
div[data-baseweb="select"] div[role="combobox"] {
    justify-content: center !important;
}
div[data-baseweb="select"] div[role="combobox"] input {
    text-align: center !important;
}

.custom-tbl { 
    width: 100%; min-width: 1100px; border-collapse: collapse; font-size: 13px; 
    font-family: "Pretendard", "Malgun Gothic", sans-serif; margin-bottom: 20px; 
    table-layout: fixed; 
}
.custom-tbl.compare-mode {
    min-width: 1900px; 
}
.ytd-wrapper { 
    display: flex; justify-content: center; width: 100%; overflow-x: auto; 
}

/* 💡 YTD 전용 테이블: 보기 좋은 간격(600px)으로 콤팩트하게 복구 */
.custom-tbl.ytd-mode {
    width: 600px !important; 
    min-width: 600px !important;
    max-width: 600px !important;
    margin: 0 auto;
}
.custom-tbl.ytd-mode td, .custom-tbl.ytd-mode th {
    padding: 6px 8px !important; 
}

.custom-tbl th { 
    text-align: center !important; padding: 10px 4px; border: 1px solid rgba(128, 128, 128, 0.2); 
    background-color: rgba(128, 128, 128, 0.05); color: inherit; white-space: nowrap; vertical-align: middle;
}
.custom-tbl td { 
    text-align: right !important; padding: 8px 10px; border: 1px solid rgba(128, 128, 128, 0.2); 
    white-space: nowrap; 
}
.custom-tbl.ytd-mode thead th {
    text-align: center !important;
}
.custom-tbl tbody td:first-child, .custom-tbl thead th.col-item { 
    width: 140px; text-align: left !important; font-weight: bold; 
    background-color: rgba(128, 128, 128, 0.02); padding-left: 10px; 
}

.pnl-container tr:has(.child-qty) { display: none; }
.pnl-container tr:has(.child-sales) { display: none; }
.pnl-container tr:has(.child-cogs) { display: none; }

#toggle-qty:checked ~ table tr:has(.child-qty) { display: table-row; }
#toggle-sales:checked ~ table tr:has(.child-sales) { display: table-row; }
#toggle-cogs:checked ~ table tr:has(.child-cogs) { display: table-row; }

.icon-qty::before, .icon-sales::before, .icon-cogs::before { content: "[+]"; color: #FF6B2C; font-weight: 900; margin-right: 6px; display: inline-block; width: 18px; }
#toggle-qty:checked ~ table .icon-qty::before { content: "[-]"; color: #242424; }
#toggle-sales:checked ~ table .icon-sales::before { content: "[-]"; color: #242424; }
#toggle-cogs:checked ~ table .icon-cogs::before { content: "[-]"; color: #242424; }

label[for="toggle-qty"], label[for="toggle-sales"], label[for="toggle-cogs"] { cursor: pointer; margin: 0; display: block; width: 100%; }
</style>
""", unsafe_allow_html=True)

def format_cell(val, is_margin):
    if pd.isna(val) or val == 0 or np.isinf(val): 
        return ""
    if is_margin: 
        return f"{val:.1f}%"
    return f"{val:,.0f}"

def clean_multiindex_html(html, is_multi=False):
    if is_multi:
        if '<th>항목</th>' in html:
            html = html.replace('<th>항목</th>', '<th class="col-item" rowspan="2" style="vertical-align: middle; border-bottom: 1px solid rgba(128,128,128,0.2);">항목</th>', 1)
            html = re.sub(r'<th[^>]*>(?:&nbsp;|\s*)</th>', '', html, count=1)
            html = re.sub(r'<th[^>]*>Unnamed[^<]*</th>', '', html)
    else:
        html = html.replace('<th>항목</th>', '<th class="col-item">항목</th>')
    return html

def render_html_table(df, mode=""):
    mode_class = " compare-mode" if mode == "compare" else (" ytd-mode" if mode == "ytd" else "")
    html = df.to_html(index=False, classes=f"custom-tbl{mode_class}", escape=False)
    html = clean_multiindex_html(html, mode in ["compare", "ytd"])
    html = html.replace("\n", "").replace("\r", "")
    wrapper_class = "ytd-wrapper" if mode == "ytd" else ""
    wrapper = f'<div class="{wrapper_class}" style="width:100%; overflow-x:auto;">{html}</div>'
    st.markdown(wrapper, unsafe_allow_html=True)

def render_pnl_table(df, mode=""):
    mode_class = " compare-mode" if mode == "compare" else (" ytd-mode" if mode == "ytd" else "")
    html = df.to_html(index=False, classes=f"custom-tbl{mode_class}", escape=False)
    html = clean_multiindex_html(html, mode in ["compare", "ytd"])
    html = html.replace("\n", "").replace("\r", "")
    wrapper_class = "pnl-container ytd-wrapper" if mode == "ytd" else "pnl-container"
    wrapper = f'<div class="{wrapper_class}" style="width:100%; overflow-x:auto;"><input type="checkbox" id="toggle-qty" style="display:none;"><input type="checkbox" id="toggle-sales" style="display:none;"><input type="checkbox" id="toggle-cogs" style="display:none;">{html}</div>'
    st.markdown(wrapper, unsafe_allow_html=True)

def render_table_unit(unit_text, is_period_compare=False):
    """표 폭에 맞춰 단위 표기를 배치한다."""
    if is_period_compare:
        style = "width: 600px; margin: 0 auto 5px auto; text-align: right;"
    else:
        style = "width: 100%; margin-bottom: 5px; text-align: right;"
    st.markdown(
        f"<div style='{style} font-size: 12px; font-weight: bold; color: #4B5563;'>{unit_text}</div>",
        unsafe_allow_html=True,
    )

def render_centered_period_selectors(months, start_key, end_key):
    """Render draft controls and return the last period confirmed by 조회."""
    initialize_period_state(st.session_state, months, start_key, end_key)
    # 중앙 열 안에서 문구와 선택 상자를 한 그룹으로 배치해 표 중심과 맞춘다.
    # 기간 설정 그룹 전체가 중앙의 기간 비교 표와 수평 중심을 맞추도록 배치한다.
    _, selector_area, _ = st.columns([0.9, 1, 1.1], gap="small")
    with selector_area:
        selector_cols = st.columns([1.25, 0.75, 0.15, 0.75, 0.55], gap="small")
    with selector_cols[0]:
        st.markdown("<div style='text-align: right; font-weight: 600; margin-top: 7px; white-space: nowrap;'>기간 설정 :</div>", unsafe_allow_html=True)
    with selector_cols[1]:
        st.selectbox("시작월", months, key=start_key, label_visibility="collapsed")
    with selector_cols[2]:
        st.markdown("<div style='text-align: center; font-weight: bold; margin-top: 5px; font-size: 16px; color: #4B5563;'>~</div>", unsafe_allow_html=True)
    with selector_cols[3]:
        st.selectbox("종료월", months, key=end_key, label_visibility="collapsed")
    with selector_cols[4]:
        query_clicked = st.button("조회", key=f"{start_key}_query", use_container_width=True)

    if query_clicked:
        result = apply_draft_period(st.session_state, months, start_key, end_key)
        if not result.applied:
            st.warning(result.error_message)

    return get_applied_period(st.session_state, months, start_key, end_key)

months = [f"{i}월" for i in range(1, 13)]


@st.cache_data(show_spinner=False)
def build_excel_template(format_items, month_labels):
    rows = []
    for row in format_items:
        if row[0] == "★손익계산서":
            rows.append(["구분", "입력 항목", "입력 키", *month_labels])
        else:
            rows.append([row[0], f"{row[1]} · {row[2]}", row[3], *([None] * 12)])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, header=False, sheet_name="Sheet1")
        from openpyxl.styles import Font, PatternFill

        worksheet = writer.sheets["Sheet1"]
        for col, width in zip(["A", "B", "C"], [18, 28, 22]):
            worksheet.column_dimensions[col].width = width
        for col in range(4, 16):
            column_letter = worksheet.cell(row=1, column=col).column_letter
            worksheet.column_dimensions[column_letter].width = 12
        header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "D2"
    return output.getvalue()

# --- 사이드바 (관리자 전용 데이터 업로드 및 양식 다운로드) ---
if st.session_state.role == "admin":
    st.sidebar.markdown("### 📁 데이터 연동 관리 (Admin)")
    
    original_format_items = [
        ('★손익계산서', '대분류', '소분류', '세분류'),
        ('Ⅰ.매출액', '1.제품 매출액', '금액(천원)', '제품매출입력'),
        ('Ⅰ.매출액', '2.반제품 매출액', '금액(천원)', '반제품매출입력'),
        ('Ⅰ.매출액', '3.상품 매출액', '금액(천원)', '상품매출입력'),
        ('Ⅰ.매출액', '4.기타 매출액', '금액(천원)', '기타매출입력'),
        ('Ⅰ.매출액', '5.판매장려금', '금액(천원)', '판매장려금입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '판매량(pcs)', 'SW수량입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '판매량(pcs)', 'BW수량입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '판매량(pcs)', 'LS수량입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '판매량(pcs)', 'FS수량입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '단가(원)', 'SW단가입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '단가(원)', 'BW단가입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '단가(원)', 'LS단가입력'),
        ('Ⅰ.매출액', '1.제품 매출액', '단가(원)', 'FS단가입력'),
        ('Ⅱ.매출원가', '1.제품 원가(투입)', '원부재료', '원부재료비입력'),
        ('Ⅱ.매출원가', '1.제품 원가(투입)', '노무비', '노무비입력'),
        ('Ⅱ.매출원가', '1.제품 원가(투입)', '외주가공비', '외주가공비입력'),
        ('Ⅱ.매출원가', '1.제품 원가(투입)', '기타경비', '기타경비입력'),
        ('Ⅱ.매출원가', '2.반제품 원가(투입)', '원부재료', '반제품_원부재료비입력'),
        ('Ⅱ.매출원가', '2.반제품 원가(투입)', '노무비', '반제품_노무비입력'),
        ('Ⅱ.매출원가', '2.반제품 원가(투입)', '외주가공비', '반제품_외주가공비입력'),
        ('Ⅱ.매출원가', '2.반제품 원가(투입)', '기타경비', '반제품_기타경비입력'),
        ('Ⅱ.매출원가', '3.반제품 매출원가(총액)', '금액(천원)', '반제품매출원가입력'),
        ('Ⅱ.매출원가', '4.상품 매출원가', '금액(천원)', '상품매출원가입력'),
        ('Ⅱ.매출원가', '5.기타 매출원가', '금액(천원)', '기타매출원가입력'),
        ('Ⅱ.매출원가', '6.표준 매출원가 차이', '금액(천원)', '표준원가차이입력'),
        ('Ⅱ.매출원가', '7.재고자산 평가손실', '금액(천원)', '재고평가손입력'),
        ('Ⅲ.매출총이익', '1.매출총이익', '금액(천원)', '매출총이익입력'), 
        ('Ⅳ.판매관리비', '1.일반관리비', '인건비', '일반관리비_인건비입력'),
        ('Ⅳ.판매관리비', '1.일반관리비', '감가상각비', '일반관리비_감가상각비입력'),
        ('Ⅳ.판매관리비', '1.일반관리비', '경상개발비', '일반관리비_경상개발비입력'),
        ('Ⅳ.판매관리비', '1.일반관리비', '수수료', '일반관리비_수수료입력'),
        ('Ⅳ.판매관리비', '1.일반관리비', '기타', '일반관리비_기타입력'),
        ('Ⅳ.판매관리비', '2.판매비', '운반비', '판매비_운반비입력'),
        ('Ⅳ.판매관리비', '2.판매비', '수수료', '판매비_수수료입력'),
        ('Ⅳ.판매관리비', '2.판매비', '브랜드사용료', '판매비_브랜드사용료입력'),
        ('Ⅳ.판매관리비', '2.판매비', '인건비', '판매비_인건비입력'),
        ('Ⅳ.판매관리비', '2.판매비', '견본비', '판…8015 tokens truncated…_total_a[i] == 0 for i in range(start_idx, end_idx + 1))
        if missing_actuals:
            st.markdown("<div style='padding: 20px; background-color: #FEE2E2; border-left: 5px solid #EF4444; border-radius: 4px; text-align: center; width: 600px; margin: 0 auto;'><h4 style='color: #B91C1C; margin: 0;'>⚠️ 실적이 없습니다</h4><p style='color: #7F1D1D; margin-top: 10px;'>선택하신 기간 중 <b>실적 데이터가 입력되지 않은 월</b>이 포함되어 비교가 불가능합니다.</p></div>", unsafe_allow_html=True)
        else:
            ytd_plan = [sum(row[start_idx:end_idx+1]) for row in plan_rows]
            ytd_actual = [sum(row[start_idx:end_idx+1]) for row in actual_rows]
            
            if ytd_plan[idx_sales] != 0:
                ytd_plan[18] = (ytd_plan[idx_cogs] / ytd_plan[idx_sales]) * 100
                ytd_plan[20] = (ytd_plan[idx_gp] / ytd_plan[idx_sales]) * 100
                ytd_plan[23] = (ytd_plan[idx_op] / ytd_plan[idx_sales]) * 100
                ytd_plan[25] = (ytd_plan[24] / ytd_plan[idx_sales]) * 100
            else:
                ytd_plan[18] = ytd_plan[20] = ytd_plan[23] = ytd_plan[25] = 0
                
            if ytd_actual[idx_sales] != 0:
                ytd_actual[18] = (ytd_actual[idx_cogs] / ytd_actual[idx_sales]) * 100
                ytd_actual[20] = (ytd_actual[idx_gp] / ytd_actual[idx_sales]) * 100
                ytd_actual[23] = (ytd_actual[idx_op] / ytd_actual[idx_sales]) * 100
                ytd_actual[25] = (ytd_actual[24] / ytd_actual[idx_sales]) * 100
            else:
                ytd_actual[18] = ytd_actual[20] = ytd_actual[23] = ytd_actual[25] = 0
                
            diff_vals = [a - p for a, p in zip(ytd_actual, ytd_plan)]
            
            ytd_tuples = [('항목', ''), (f'{selected_start_m}~{selected_end_m} 누계', '계획'), (f'{selected_start_m}~{selected_end_m} 누계', '실적'), (f'{selected_start_m}~{selected_end_m} 누계', '차이(실적-계획)')]
            ytd_rows_data = []
            
            for i, item in enumerate(items):
                is_ratio = ('율' in str(item) or '률' in str(item))
                if pd.isna(diff_vals[i]) or np.isinf(diff_vals[i]): diff_str = ""
                elif is_ratio: diff_str = f"{diff_vals[i]:+.1f}%p" if diff_vals[i] != 0 else "0.0%p"
                else: diff_str = f"{diff_vals[i]:+,.0f}" if diff_vals[i] != 0 else "0"
                ytd_rows_data.append([item, format_cell(ytd_plan[i], is_ratio), format_cell(ytd_actual[i], is_ratio), diff_str])
                
            df_ytd = pd.DataFrame(ytd_rows_data, columns=pd.MultiIndex.from_tuples(ytd_tuples))
            render_pnl_table(df_ytd, "ytd")

elif view_mode == "계획/실적 비교":
    tuples = [('항목', '')]
    for m in months: tuples.extend([(m, '계획'), (m, '실적')])
    tuples.extend([('합계', '계획'), ('합계', '실적')])
    c_rows = []
    for i, item in enumerate(items):
        r_data = [item]
        for m_idx in range(12): r_data.extend([plan_rows[i][m_idx], actual_rows[i][m_idx]])
        r_data.extend([plan_sums[i], actual_sums[i]])
        c_rows.append(r_data)
    df_table = pd.DataFrame(c_rows, columns=pd.MultiIndex.from_tuples(tuples))
    for col in df_table.columns:
        if col != ('항목', ''): df_table[col] = df_table.apply(lambda row: format_cell(row[col], '율' in str(row[('항목', '')]) or '률' in str(row[('항목', '')])), axis=1)
    render_pnl_table(df_table, "compare")
else:
    df_table = pd.DataFrame({'항목': items})
    for i, month in enumerate(months): df_table[month] = [row[i] for row in actual_rows]
    df_table['합계'] = actual_sums
    for col in df_table.columns:
        if col != '항목': df_table[col] = df_table.apply(lambda row: format_cell(row[col], '율' in str(row['항목']) or '률' in str(row['항목'])), axis=1)
    render_pnl_table(df_table, "")

# 8. 제품/반제품 매출원가 내역
st.markdown("---")
st.markdown("##### 🔍 제품/반제품(FS) 매출원가 내역")

cogs_prod_input_sum_a = cogs_rm_a + cogs_lb_a + cogs_os_a + cogs_oh_a
cogs_semi_input_sum_a = cogs_semi_rm_a + cogs_semi_lb_a + cogs_semi_os_a + cogs_semi_oh_a
comb_input_sum_a = cogs_prod_input_sum_a + cogs_semi_input_sum_a

cogs_prod_input_sum_p = cogs_rm_p + cogs_lb_p + cogs_os_p + cogs_oh_p
cogs_semi_input_sum_p = cogs_semi_rm_p + cogs_semi_lb_p + cogs_semi_os_p + cogs_semi_oh_p
comb_input_sum_p = cogs_prod_input_sum_p + cogs_semi_input_sum_p

target_comb_cogs_a = cogs_prod_a + cogs_semi_a
target_comb_cogs_p = cogs_prod_p + cogs_semi_p

mismatched_months_a = [f"{i+1}월" for i in range(12) if abs(target_comb_cogs_a[i] - comb_input_sum_a[i]) >= 1.0]
mismatched_months_p = [f"{i+1}월" for i in range(12) if abs(target_comb_cogs_p[i] - comb_input_sum_p[i]) >= 1.0]

if mismatched_months_a or mismatched_months_p:
    st.markdown("<div style='padding: 15px; background-color: #FEE2E2; border-left: 5px solid #EF4444; border-radius: 4px; margin-bottom: 15px;'>", unsafe_allow_html=True)
    st.markdown("<p style='color: #B91C1C; font-weight: bold; margin: 0;'>⚠️ [합계 오류] 손익계산서 상 제품/반제품 매출원가와 하단 내역의 합계가 일치하지 않습니다.</p>", unsafe_allow_html=True)
    if mismatched_months_p:
        st.markdown(f"<p style='color: #7F1D1D; margin: 5px 0 0 0;'>• <b>계획(Plan) 점검 필요:</b> {', '.join(mismatched_months_p)}</p>", unsafe_allow_html=True)
    if mismatched_months_a:
        st.markdown(f"<p style='color: #7F1D1D; margin: 5px 0 0 0;'>• <b>실적(Actual) 점검 필요:</b> {', '.join(mismatched_months_a)}</p>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown("<div style='text-align: right; font-size: 12px; font-weight: bold; color: #4B5563; margin-bottom: 5px;'>(단위: 백만원, %)</div>", unsafe_allow_html=True)

comb_rm_a = cogs_rm_a + cogs_semi_rm_a
comb_lb_a = cogs_lb_a + cogs_semi_lb_a
comb_os_a = cogs_os_a + cogs_semi_os_a
comb_oh_a = cogs_oh_a + cogs_semi_oh_a

cogs_items = ['원부재료', '노무비', '외주가공비', '기타경비', '합계'] 
cogs_rows_a = [comb_rm_a, comb_lb_a, comb_os_a, comb_oh_a, comb_input_sum_a]
cogs_sums_a = [sum(row) for row in cogs_rows_a]

sales_denom_a = sales_total_a
sales_denom_sum_a = sum(sales_denom_a)

tuples_cogs = [('항목', '')]
for m in months: tuples_cogs.extend([(m, '실적금액'), (m, '매출비율')])
tuples_cogs.extend([('합계', '실적금액'), ('합계', '매출비율')])

combined_rows_cogs = []
for i, item in enumerate(cogs_items):
    row_data = [item]
    for m_idx in range(12):
        amt = cogs_rows_a[i][m_idx]
        ratio = (amt / sales_denom_a[m_idx]) * 100 if sales_denom_a[m_idx] != 0 else 0
        row_data.extend([amt, format_cell(ratio, True) if amt != 0 else ""])
    sum_amt = cogs_sums_a[i]
    sum_ratio = (sum_amt / sales_denom_sum_a) * 100 if sales_denom_sum_a != 0 else 0
    row_data.extend([sum_amt, format_cell(sum_ratio, True) if sum_amt != 0 else ""])
    combined_rows_cogs.append(row_data)

df_cogs = pd.DataFrame(combined_rows_cogs, columns=pd.MultiIndex.from_tuples(tuples_cogs))
for col in df_cogs.columns:
    if col[1] == '실적금액': df_cogs[col] = df_cogs[col].apply(lambda x: format_cell(x, False))
render_html_table(df_cogs, "compare")
st.markdown("<br>", unsafe_allow_html=True)

# 9. 판매관리비 명세서
st.markdown("---")
sga_mode = st.session_state.get("sga_toggle", "실적만 보기")
st.markdown("##### 🔍 판매관리비 명세서")
st.radio("판관비 표시 기준 선택", ["실적만 보기", "계획/실적 비교", "기간 설정 비교"], horizontal=True, label_visibility="collapsed", key="sga_toggle")

if sga_mode == "기간 설정 비교":
    selected_st_sga, selected_ed_sga = render_centered_period_selectors(months, "st_sga", "ed_sga")
    start_idx_sga = months.index(selected_st_sga)
    end_idx_sga = months.index(selected_ed_sga)
else:
    start_idx_sga = 0
    end_idx_sga = 0
    selected_st_sga = "1월"
    selected_ed_sga = "1월"

view_mode_sga = st.session_state["sga_toggle"]

render_table_unit("(단위: 백만원)", view_mode_sga == "기간 설정 비교")

sga_items = [
    '【 일반관리비 소계 】', ' - 인건비', ' - 감가상각비', ' - 경상개발비', ' - 수수료', ' - 기타',
    '【 판매비 소계 】', ' - 운반비', ' - 수수료', ' - 브랜드사용료', ' - 인건비', ' - 견본비', ' - 대손상각', ' - 잡비', ' - 기타',
    '▶ 판관비 총계'
]

sga_actual_rows = [
    adm_total_a, adm_labor_a, adm_depr_a, adm_rnd_a, adm_fee_a, adm_etc_a,
    sel_total_a, sel_trans_a, sel_fee_a, sel_brand_a, sel_labor_a, sel_sample_a, sel_bad_a, sel_misc_a, sel_etc_a,
    sga_total_a
]

sga_plan_rows = [
    adm_total_p, adm_labor_p, adm_depr_p, adm_rnd_p, adm_fee_p, adm_etc_p,
    sel_total_p, sel_trans_p, sel_fee_p, sel_brand_p, sel_labor_p, sel_sample_p, sel_bad_p, sel_misc_p, sel_etc_p,
    sga_total_p
]

sga_actual_sums = [sum(row) for row in sga_actual_rows]
sga_plan_sums = [sum(row) for row in sga_plan_rows]

if view_mode_sga == "기간 설정 비교":
    if start_idx_sga > end_idx_sga:
        st.markdown("<div style='padding: 20px; background-color: #FEE2E2; border-left: 5px solid #EF4444; border-radius: 4px; text-align: center; width: 600px; margin: 0 auto;'><h4 style='color: #B91C1C; margin: 0;'>⚠️ 기간 설정 오류</h4><p style='color: #7F1D1D; margin-top: 10px;'>시작월이 종료월보다 이후일 수 없습니다.</p></div>", unsafe_allow_html=True)
    else:
        missing_actuals = any(sales_total_a[i] == 0 for i in range(start_idx_sga, end_idx_sga + 1))
        if missing_actuals:
            st.markdown("<div style='padding: 20px; background-color: #FEE2E2; border-left: 5px solid #EF4444; border-radius: 4px; text-align: center; width: 600px; margin: 0 auto;'><h4 style='color: #B91C1C; margin: 0;'>⚠️ 실적이 없습니다</h4><p style='color: #7F1D1D; margin-top: 10px;'>선택하신 기간 중 <b>실적 데이터가 입력되지 않은 월</b>이 포함되어 비교가 불가능합니다.</p></div>", unsafe_allow_html=True)
        else:
            ytd_plan_sga = [sum(row[start_idx_sga:end_idx_sga+1]) for row in sga_plan_rows]
            ytd_actual_sga = [sum(row[start_idx_sga:end_idx_sga+1]) for row in sga_actual_rows]
            diff_vals_sga = [a - p for a, p in zip(ytd_actual_sga, ytd_plan_sga)]
            
            ytd_tuples_sga = [('항목', ''), (f'{selected_st_sga}~{selected_ed_sga} 누계', '계획'), (f'{selected_st_sga}~{selected_ed_sga} 누계', '실적'), (f'{selected_st_sga}~{selected_ed_sga} 누계', '차이(실적-계획)')]
            ytd_rows_data_sga = []
            for i, item in enumerate(sga_items):
                diff_str = f"{diff_vals_sga[i]:+,.0f}" if diff_vals_sga[i] != 0 else "0"
                ytd_rows_data_sga.append([item, format_cell(ytd_plan_sga[i], False), format_cell(ytd_actual_sga[i], False), diff_str])
            
            df_ytd_sga = pd.DataFrame(ytd_rows_data_sga, columns=pd.MultiIndex.from_tuples(ytd_tuples_sga))
            render_html_table(df_ytd_sga, "ytd")
            
elif view_mode_sga == "계획/실적 비교":
    tuples_sga = [('항목', '')]
    for m in months: tuples_sga.extend([(m, '계획'), (m, '실적')])
    tuples_sga.extend([('합계', '계획'), ('합계', '실적')])
    combined_rows_sga = []
    for i, item in enumerate(sga_items):
        row_data = [item]
        for m_idx in range(12): row_data.extend([sga_plan_rows[i][m_idx], sga_actual_rows[i][m_idx]])
        row_data.extend([sga_plan_sums[i], sga_actual_sums[i]])
        combined_rows_sga.append(row_data)
    df_sga_comp = pd.DataFrame(combined_rows_sga, columns=pd.MultiIndex.from_tuples(tuples_sga))
    for col in df_sga_comp.columns:
        if col != ('항목', ''): df_sga_comp[col] = df_sga_comp[col].apply(lambda x: format_cell(x, False))
    render_html_table(df_sga_comp, "compare")
else:
    df_sga = pd.DataFrame({'항목': sga_items})
    for i, month in enumerate(months): df_sga[month] = [row[i] for row in sga_actual_rows]
    df_sga['합계'] = sga_actual_sums
    for col in df_sga.columns:
        if col != '항목': df_sga[col] = df_sga[col].apply(lambda x: format_cell(x, False))
    render_html_table(df_sga, "")

# 10. Item별 구분손익
st.markdown("---")
type_mode = st.session_state.get("type_toggle", "실적만 보기")
st.markdown("##### 🔍 Item별 구분손익")
st.radio("구분손익 표시 기준 선택", ["실적만 보기", "계획/실적 비교", "기간 설정 비교"], horizontal=True, label_visibility="collapsed", key="type_toggle")

if type_mode == "기간 설정 비교":
    selected_st_type, selected_ed_type = render_centered_period_selectors(months, "st_type", "ed_type")
    start_idx_type = months.index(selected_st_type)
    end_idx_type = months.index(selected_ed_type)
else:
    start_idx_type = 0
    end_idx_type = 0
    selected_st_type = "1월"
    selected_ed_type = "1월"

view_mode_type = st.session_state["type_toggle"]

def build_type_pnl(qty_a, qty_p, price_a, price_p, cogs_a, cogs_p, sga_a, sga_p):
    sales_a = (qty_a * price_a) / 1000000.0
    sales_p = (qty_p * price_p) / 1000000.0
    
    gp_a, gp_p = sales_a - cogs_a, sales_p - cogs_p
    op_a, op_p = gp_a - sga_a, gp_p - sga_p
    
    cr_a = np.zeros(12); cr_p = np.zeros(12)
    gpr_a = np.zeros(12); gpr_p = np.zeros(12)
    opr_a = np.zeros(12); opr_p = np.zeros(12)
    for i in range(12):
        if sales_a[i] != 0:
            cr_a[i] = (cogs_a[i] / sales_a[i]) * 100
            gpr_a[i] = (gp_a[i] / sales_a[i]) * 100
            opr_a[i] = (op_a[i] / sales_a[i]) * 100
        if sales_p[i] != 0:
            cr_p[i] = (cogs_p[i] / sales_p[i]) * 100
            gpr_p[i] = (gp_p[i] / sales_p[i]) * 100
            opr_p[i] = (op_p[i] / sales_p[i]) * 100

    rows_a = [sales_a, qty_a, price_a, cogs_a, cr_a, gp_a, gpr_a, sga_a, op_a, opr_a]
    rows_p = [sales_p, qty_p, price_p, cogs_p, cr_p, gp_p, gpr_p, sga_p, op_p, opr_p]

    s_sales_a, s_sales_p = sum(sales_a), sum(sales_p)
    avg_price_a = (s_sales_a * 1000000.0) / sum(qty_a) if sum(qty_a) else 0
    avg_price_p = (s_sales_p * 1000000.0) / sum(qty_p) if sum(qty_p) else 0
    
    sums_a = [s_sales_a, sum(qty_a), avg_price_a, sum(cogs_a), (sum(cogs_a)/s_sales_a*100) if s_sales_a else 0, sum(gp_a), (sum(gp_a)/s_sales_a*100) if s_sales_a else 0, sum(sga_a), sum(op_a), (sum(op_a)/s_sales_a*100) if s_sales_a else 0]
    sums_p = [s_sales_p, sum(qty_p), avg_price_p, sum(cogs_p), (sum(cogs_p)/s_sales_p*100) if s_sales_p else 0, sum(gp_p), (sum(gp_p)/s_sales_p*100) if s_sales_p else 0, sum(sga_p), sum(op_p), (sum(op_p)/s_sales_p*100) if s_sales_p else 0]
    return rows_a, sums_a, rows_p, sums_p

type_items = ['매출액', '매출수량(pcs)', '단가(원)', '매출원가', '매출원가율', '매출총이익', '매출총이익률', '판관비', '영업이익', '영업이익률']

sw_rows_a, sw_sums_a, sw_rows_p, sw_sums_p = build_type_pnl(qty_sw_a, qty_sw_p, price_sw_a, price_sw_p, cogs_sw_a, cogs_sw_p, sga_sw_a, sga_sw_p)
bw_rows_a, bw_sums_a, bw_rows_p, bw_sums_p = build_type_pnl(qty_bw_a, qty_bw_p, price_bw_a, price_bw_p, cogs_bw_a, cogs_bw_p, sga_bw_a, sga_bw_p)

def render_type_table(rows_a, sums_a, rows_p, sums_p, view_mode, st_idx=0, ed_idx=0, st_month="", ed_month=""):
    if view_mode == "기간 설정 비교":
        if st_idx > ed_idx:
            st.markdown("<div style='padding: 20px; background-color: #FEE2E2; border-left: 5px solid #EF4444; border-radius: 4px; text-align: center; width: 600px; margin: 0 auto;'><h4 style='color: #B91C1C; margin: 0;'>⚠️ 기간 설정 오류</h4><p style='color: #7F1D1D; margin-top: 10px;'>시작월이 종료월보다 이후일 수 없습니다.</p></div>", unsafe_allow_html=True)
            return
            
        missing_actuals = any(rows_a[0][i] == 0 for i in range(st_idx, ed_idx + 1))
        if missing_actuals:
            st.markdown("<div style='padding: 20px; background-color: #FEE2E2; border-left: 5px solid #EF4444; border-radius: 4px; text-align: center; width: 600px; margin: 0 auto;'><h4 style='color: #B91C1C; margin: 0;'>⚠️ 실적이 없습니다</h4><p style='color: #7F1D1D; margin-top: 10px;'>선택하신 기간 중 <b>실적 데이터가 입력되지 않은 월</b>이 포함되어 비교가 불가능합니다.</p></div>", unsafe_allow_html=True)
            return
            
        ytd_plan = [sum(r[st_idx:ed_idx+1]) for r in rows_p]
        ytd_actual = [sum(r[st_idx:ed_idx+1]) for r in rows_a]
        
        for ytd_arr in [ytd_plan, ytd_actual]:
            if ytd_arr[0] != 0:
                ytd_arr[4] = (ytd_arr[3] / ytd_arr[0]) * 100
                ytd_arr[6] = (ytd_arr[5] / ytd_arr[0]) * 100
                ytd_arr[9] = (ytd_arr[8] / ytd_arr[0]) * 100
                ytd_arr[2] = (ytd_arr[0] * 1000000.0) / ytd_arr[1] if ytd_arr[1] != 0 else 0
            else:
                ytd_arr[4] = ytd_arr[6] = ytd_arr[9] = ytd_arr[2] = 0
                
        diff_vals = [a - p for a, p in zip(ytd_actual, ytd_plan)]
        tuples = [('항목', ''), (f'{st_month}~{ed_month} 누계', '계획'), (f'{st_month}~{ed_month} 누계', '실적'), (f'{st_month}~{ed_month} 누계', '차이(실적-계획)')]
        c_rows = []
        for i, item in enumerate(type_items):
            is_ratio = ('율' in str(item) or '률' in str(item))
            diff_str = ""
            if pd.isna(diff_vals[i]) or np.isinf(diff_vals[i]): diff_str = ""
            elif is_ratio: diff_str = f"{diff_vals[i]:+.1f}%p" if diff_vals[i] != 0 else "0.0%p"
            else: diff_str = f"{diff_vals[i]:+,.0f}" if diff_vals[i] != 0 else "0"
            c_rows.append([item, format_cell(ytd_plan[i], is_ratio), format_cell(ytd_actual[i], is_ratio), diff_str])
            
        df = pd.DataFrame(c_rows, columns=pd.MultiIndex.from_tuples(tuples))
        render_html_table(df, "ytd")
        
    elif view_mode == "계획/실적 비교":
        tuples = [('항목', '')]
        for m in months: tuples.extend([(m, '계획'), (m, '실적')])
        tuples.extend([('합계', '계획'), ('합계', '실적')])
        c_rows = []
        for i, item in enumerate(type_items):
            r_data = [item]
            for m_idx in range(12): r_data.extend([rows_p[i][m_idx], rows_a[i][m_idx]])
            r_data.extend([sums_p[i], sums_a[i]])
            c_rows.append(r_data)
        df = pd.DataFrame(c_rows, columns=pd.MultiIndex.from_tuples(tuples))
        for col in df.columns:
            if col != ('항목', ''): df[col] = df.apply(lambda row: format_cell(row[col], '율' in str(row[('항목', '')]) or '률' in str(row[('항목', '')])), axis=1)
        render_html_table(df, "compare")
    else:
        df = pd.DataFrame({'항목': type_items})
        for i, m in enumerate(months): df[m] = [r[i] for r in rows_a]
        df['합계'] = sums_a
        for col in df.columns:
            if col != '항목': df[col] = df.apply(lambda row: format_cell(row[col], '율' in str(row['항목']) or '률' in str(row['항목'])), axis=1)
        render_html_table(df, "")


tab1, tab2 = st.tabs(["8인치 SW", "8인치 BW"])
with tab1: 
    st.markdown("**■ 8인치 SW 손익 명세**")
    render_table_unit("(단위: 백만원, pcs, 원, %)", view_mode_type == "기간 설정 비교")
    render_type_table(sw_rows_a, sw_sums_a, sw_rows_p, sw_sums_p, view_mode_type, start_idx_type, end_idx_type, selected_st_type, selected_ed_type)
with tab2: 
    st.markdown("**■ 8인치 BW 손익 명세**")
    render_table_unit("(단위: 백만원, pcs, 원, %)", view_mode_type == "기간 설정 비교")
    render_type_table(bw_rows_a, bw_sums_a, bw_rows_p, bw_sums_p, view_mode_type, start_idx_type, end_idx_type, selected_st_type, selected_ed_type)

# Item별 구분손익 표 하단 여백
st.markdown("<div style='height: 64px;'></div>", unsafe_allow_html=True)

