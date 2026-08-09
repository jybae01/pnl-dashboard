from __future__ import annotations

import json
import re
import shutil
import tempfile
import zipfile
from collections import Counter
from datetime import date, datetime
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from .formula import FormulaError, RangeValue, col_name, col_number, evaluate, normalize_ref, translate_formula

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NAMESPACES = {
    "": NS,
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "x15": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main",
    "x15ac": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac",
    "x16r2": "http://schemas.microsoft.com/office/spreadsheetml/2015/02/main",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
}
for _prefix, _uri in NAMESPACES.items():
    ET.register_namespace(_prefix, _uri)
Q = lambda name: f"{{{NS}}}{name}"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
AUDIT_SHEET_NAME = "ì…ë ¥ë°˜ì˜ë‚´ì—­"
MONTH_COLUMNS = {month: chr(ord("E") + month - 1) for month in range(1, 13)}
FORMULA_REFERENCE_RE = re.compile(
    r"(?<![A-Z0-9_])(?P<start>\$?[A-Z]{1,3}\$?\d+)"
    r"(?::(?P<end>\$?[A-Z]{1,3}\$?\d+))?",
    re.IGNORECASE,
)
PERIOD_TYPE_ALIASES = {
    "actual": "ì‹¤ì ",
    "actuals": "ì‹¤ì ",
    "ì‹¤ì ": "ì‹¤ì ",
    "estimate": "ì¶”ì •",
    "estimated": "ì¶”ì •",
    "forecast": "ì¶”ì •",
    "ì¶”ì •": "ì¶”ì •",
    "plan": "ê³„íš",
    "planned": "ê³„íš",
    "ê³„íš": "ê³„íš",
}


def _normalize_period_type(value: Any) -> str:
    text = str(value or "").strip()
    return PERIOD_TYPE_ALIASES.get(text.lower(), text)


def extract_period_types(path: str | Path) -> dict[str, str]:
    """Read the 1~12ì›” status row (Data!E3:P3) from an uploaded workbook.

    Golden Models keep the month columns fixed; only the status in row 3
    changes between ì‹¤ì /ì¶”ì •/ê³„íš. Returning all twelve keys keeps the
    metadata shape stable even when a workbook leaves a status blank.
    """
    path = Path(path)
    statuses: dict[str, str] = {str(month): "" for month in range(1, 13)}

    # openpyxl understands inline strings and date-formatted cells that are
    # otherwise awkward to interpret from OOXML alone. It is already a
    # runtime dependency because the analysis export uses it.
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        worksheet = next(
            (workbook[name] for name in workbook.sheetnames if name.lower() == "data"),
            workbook[workbook.sheetnames[0]],
        )
        for month, column in MONTH_COLUMNS.items():
            statuses[str(month)] = _normalize_period_type(worksheet[f"{column}3"].value)
        workbook.close()
        return statuses
    except Exception:
        # Fall back to the project's OOXML reader for workbooks that contain
        # an Excel extension unsupported by openpyxl.
        workbook = GoldenWorkbook(path)
        for month, column in MONTH_COLUMNS.items():
            statuses[str(month)] = _normalize_period_type(workbook.raw_value(f"{column}3"))
        return statuses


def summarize_period_types(period_types: dict[str, str] | None) -> str:
    """Format contiguous monthly statuses for a compact read-only UI label."""
    values = [_normalize_period_type((period_types or {}).get(str(month))) for month in range(1, 13)]
    if not any(values):
        return "ë¯¸í™•ì¸"
    segments: list[str] = []
    start = 1
    current = values[0] or "ë¯¸ì§€ì •"
    for month in range(2, 13):
        value = values[month - 1] or "ë¯¸ì§€ì •"
        if value == current:
            continue
        end = month - 1
        label = f"{start}ì›”" if start == end else f"{start}~{end}ì›”"
        segments.append(f"{label} {current}")
        start, current = month, value
    end = 12
    label = f"{start}ì›”" if start == end else f"{start}~{end}ì›”"
    segments.append(f"{label} {current}")
    return " / ".join(segments)


def _year_from_value(value: Any) -> int | None:
    if isinstance(value, (datetime, date)):
        return value.year if 1900 <= value.year <= 2100 else None
    if isinstance(value, str):
        match = re.search(r"(?<!\d)(19\d{2}|20\d{2})(?!\d)", value)
        return int(match.group(1)) if match else None
    return None


def infer_workbook_year(path: str | Path, fallback_year: int | None = None) -> int:
    """Infer the model year from month/date/header cells.

    The upload form deliberately has no year input. We first inspect explicit
    four-digit years, date cells, and date-formatted numeric headers, giving
    the top rows of the Data sheet priority. If no year is discoverable, the
    documented fallback is ``fallback_year``; when it is omitted, the current
    calendar year is used.
    """
    path = Path(path)
    fallback = int(fallback_year or date.today().year)
    candidates: list[tuple[int, int]] = []
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        sheets = list(workbook.worksheets)
        sheets.sort(key=lambda sheet: 0 if sheet.title.lower() == "data" else 1)
        for worksheet in sheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    year = _year_from_value(value)
                    if year is None and isinstance(value, (int, float)) and not isinstance(value, bool):
                        number_format = str(cell.number_format or "").lower()
                        if 1900 <= value <= 2100 and ("yy" in number_format or cell.row <= 3):
                            year = int(value)
                    if year is None:
                        continue
                    score = 1
                    if worksheet.title.lower() == "data":
                        score += 3
                    if cell.row <= 5:
                        score += 3
                    if cell.column in range(5, 17):
                        score += 1
                    candidates.append((year, score))
        workbook.close()
    except Exception:
        # The custom reader still handles the simple inline-string/date labels
        # used by Golden Models when openpyxl cannot load an extension.
        try:
            workbook = GoldenWorkbook(path)
            for address in workbook.cells:
                value = workbook.raw_value(address)
                year = _year_from_value(value)
                row_match = re.search(r"\d+", address)
                row = int(row_match.group(0)) if row_match else 9999
                if year is None and isinstance(value, (int, float)) and 1900 <= value <= 2100 and row <= 3:
                    year = int(value)
                if year is not None:
                    candidates.append((year, 4 if row <= 5 else 1))
        except Exception:
            candidates = []

    if candidates:
        scores = Counter()
        counts = Counter()
        for year, score in candidates:
            scores[year] += score
            counts[year] += 1
        return max(scores, key=lambda year: (scores[year], counts[year], -abs(year - fallback)))
    # Fallback rule: use an explicitly supplied existing ê¸°ì¤€ì—°ë„ when one is
    # available; otherwise use today's calendar year for a new upload.
    return fallback


def serialize_xml(
    root: ET.Element,
    required_prefixes: tuple[str, ...] = (),
    default_namespace: str | None = None,
) -> bytes:
    """Serialize without breaking mc:Ignorable namespace declarations.

    ElementTree omits namespace declarations that occur only inside the text
    value of mc:Ignorable. Desktop Excel treats that as workbook corruption.
    """
    text = ET.tostring(root, encoding="unicode", xml_declaration=False)
    if default_namespace:
        declaration = re.search(
            rf'xmlns:(?P<prefix>[A-Za-z_][\w.-]*)="{re.escape(default_namespace)}"',
            text,
        )
        if declaration:
            prefix = declaration.group("prefix")
            text = text.replace(f"<{prefix}:", "<").replace(f"</{prefix}:", "</")
            text = text.replace(
                f'xmlns:{prefix}="{default_namespace}"',
                f'xmlns="{default_namespace}"',
            )
    opening_end = text.find(">")
    opening = text[:opening_end]
    additions = []
    for prefix in required_prefixes:
        if f"xmlns:{prefix}=" not in opening:
            additions.append(f' xmlns:{prefix}="{NAMESPACES[prefix]}"')
    if additions:
        text = text[:opening_end] + "".join(additions) + text[opening_end:]
    return ("<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>\r\n" + text).encode("utf-8")


@dataclass
class ChangeLog:
    cell: str
    old_value: Any
    new_value: Any
    source: str
    reason: str
    formula_overwritten: bool = False


@dataclass(frozen=True)
class FormulaFallback:
    address: str
    formula: str
    exception_type: str
    exception_message: str
    cached_value: Any


class GoldenWorkbook:
    """OOXML adapter that changes only configured cells and preserves the package."""

    def __init__(self, path: str | Path, sheet_xml: str | None = None):
        self.path = Path(path)
        with zipfile.ZipFile(self.path) as archive:
            self.entries = {name: archive.read(name) for name in archive.namelist()}
        if sheet_xml is None:
            sheet_xml = self._default_model_sheet_xml()
        if sheet_xml not in self.entries:
            raise ValueError(f"ì›Œí¬ë¶ì—ì„œ ëª¨í˜• ì‹œíŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {sheet_xml}")
        self.sheet_xml = sheet_xml
        self.shared_strings = self._read_shared_strings()
        self.root = ET.fromstring(self.entries[self.sheet_xml])
        self.cells = {cell.attrib["r"].upper(): cell for cell in self.root.iter(Q("c"))}
        self.formulas: dict[str, str] = {}
        self._expand_formulas()
        self.original_formulas = dict(self.formulas)
        self.overrides: dict[str, Any] = {}
        self.cache: dict[str, Any] = {}
        self.stack: set[str] = set()
        self.formula_successes: set[str] = set()
        self.formula_fallbacks: dict[str, FormulaFallback] = {}
        self._precedent_cache: dict[str, tuple[str, ...]] = {}
        self.log: list[ChangeLog] = []

    def _default_model_sheet_xml(self) -> str:
        """Prefer the approved Data sheet even when a README sheet comes first."""
        for alias in ("Data", "DATA", "ì›ë³¸_Data", "ì›ë³¸ë°ì´í„°"):
            resolved = self._worksheet_target(self.entries, alias)
            if resolved is not None:
                return resolved[0]
        return "xl/worksheets/sheet1.xml"

    def _read_shared_strings(self) -> list[str]:
        raw = self.entries.get("xl/sharedStrings.xml")
        if not raw: return []
        root = ET.fromstring(raw)
        return ["".join(node.text or "" for node in item.iter(Q("t"))) for item in root.findall(Q("si"))]

    def _expand_formulas(self) -> None:
        anchors: dict[str, tuple[str, str]] = {}
        pending: list[tuple[str, str]] = []
        for addr, cell in self.cells.items():
            node = cell.find(Q("f"))
            if node is None: continue
            text = (node.text or "").strip()
            si = node.attrib.get("si")
            if node.attrib.get("t") == "shared" and si:
                if text:
                    anchors[si] = (addr, text)
                else:
                    pending.append((addr, si))
            if text:
                self.formulas[addr] = "=" + text.lstrip("=")
        for addr, si in pending:
            if si not in anchors:
                raise ValueError(f"Shared formula anchor missing: {addr}, si={si}")
            source, formula = anchors[si]
            self.formulas[addr] = "=" + translate_formula(formula, source, addr).lstrip("=")

    def raw_value(self, addr: str) -> Any:
        addr = normalize_ref(addr)
        if addr in self.overrides: return self.overrides[addr]
        cell = self.cells.get(addr)
        if cell is None: return None
        kind = cell.attrib.get("t")
        if kind == "inlineStr":
            return "".join(node.text or "" for node in cell.iter(Q("t")))
        value = cell.find(Q("v"))
        if value is None or value.text is None: return None
        if kind == "s": return self.shared_strings[int(value.text)]
        if kind == "b": return value.text == "1"
        if kind in ("str", "e"): return value.text
        try: return float(value.text)
        except ValueError: return value.text

    def value(self, addr: str) -> Any:
        addr = normalize_ref(addr)
        if addr in self.overrides: return self.overrides[addr]
        if addr in self.cache: return self.cache[addr]
        formula = self.formulas.get(addr)
        if not formula: return self.raw_value(addr)
        if addr in self.stack: raise FormulaError(f"Circular reference at {addr}")
        self.stack.add(addr)
        try:
            result = evaluate(formula, self.value, self.range_value)
        except (FormulaError, ZeroDivisionError, TypeError, ValueError) as exc:
            result = self.raw_value(addr)
            self.formula_successes.discard(addr)
            self.formula_fallbacks[addr] = FormulaFallback(
                address=addr,
                formula=formula,
                exception_type=type(exc).__name__,
                exception_message=str(exc),
                cached_value=result,
            )
        else:
            self.formula_fallbacks.pop(addr, None)
            self.formula_successes.add(addr)
        finally:
            self.stack.remove(addr)
        self.cache[addr] = result
        return reã~¶¶‰Ëkºwµç@€€€€€€€€€€¡…¹•1½œ 4(€€€€€€€€€€€€€€€€€€€•±°õÍÑÈ¡É•½É‘l‰•±°‰t¤°4(€€€€€€€€€€€€€€€€€€€½±‘}Ù…±Õ”õÉ•½É¹•Ğ ‰½±‘}Ù…±Õ”ˆ¤°4(€€€€€€€€€€€€€€€€€€€¹•İ}Ù…±Õ”õÉ•½É¹•Ğ ‰¹•İ}Ù…±Õ”ˆ¤°4(€€€€€€€€€€€€€€€€€€€Í½ÕÉ”õÍÑÈ¡É•½É‘l‰Í½ÕÉ”‰t¤°4(€€€€€€€€€€€€€€€€€€€É•…Í½¸õÍÑÈ¡É•½É¹•Ğ ‰É•…Í½¸ˆ°€ˆˆ¤¤°4(€€€€€€€€€€€€€€€€€€€™½ÉµÕ±…}½Ù•ÉİÉ¥ÑÑ•¸õ‰½½°¡É•½É¹•Ğ ‰™½ÉµÕ±…}½Ù•ÉİÉ¥ÑÑ•¸ˆ°…±Í”¤¤°4(€€€€€€€€€€€€€€€€¤4(€€€€€€€€€€€€¤4(4(€€€ÍÑ…Ñ¥µ•Ñ¡½4(€€€‘•˜}…Õ‘¥Ñ}…Ñ•½Éä¡Í½ÕÉ”èÍÑÈ¤€´øÍÑÈè4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰Í…±•Ì¸ˆ¤è4(€€€€€€€€€€€É•ÑÕÉ¸€‹¶2C®ˆ4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰ÁÉ½‘ÕÑ¥½¸¸ˆ¤è4(€€€€€€€€€€€É•ÑÕÉ¸€‹²w²
Àˆ4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰µ´¸ˆ¤è4(€€€€€€€€€€€É•ÑÕÉ¸€‰54£²rƒ²²
³ªâ$¤ˆ4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰É…İ}µ…Ñ•É¥…°¹ÁÕÉ¡…Í•}•ÍÑ¥µ…Ñ”¸ˆ¤è4(€€€€€€€€€€€É•ÑÕÉ¸€‹²nC²z³®0ƒ¶"³²z®æˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰µ…¹Õ™…ÑÕÉ¥¹}…‘©ÕÍÑµ•¹Ğˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹²‚s²†ÃªÊ÷®æˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰Í…}…‘©ÕÍÑµ•¹Ğˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹¶2CªÒ®æˆ4(€€€€€€€É•ÑÕÉ¸€‹®“²Ús²nCªÂ ˆ4(4(€€€‘•˜}É½İ}±…‰•°¡Í•±˜°É½Üè¥¹Ğ¤€´øÍÑÈè4(€€€€€€€™½È½±Õµ¸¥¸€ ‰ˆ°€‰ˆ°€‰ˆ°€‰ˆ¤è4(€€€€€€€€€€€Ù…±Õ”€ôÍ•±˜¹É…İ}Ù…±Õ”¡˜‰í½±Õµ¹õíÉ½İôˆ¤4(€€€€€€€€€€€¥˜Ù…±Õ”¹½Ğ¥¸€¡9½¹”°€ˆˆ¤è4(€€€€€€€€€€€€€€€É•ÑÕÉ¸ÍÑÈ¡Ù…±Õ”¤¹ÍÑÉ¥À ¤4(€€€€€€€É•ÑÕÉ¸˜‰íÉ½İ÷¶Z$ˆ4(4(€€€‘•˜}Í…}Í•Ñ¥½¸¡Í•±˜°É½Üè¥¹Ğ¤€´øÍÑÈè4(€€€€€€€€ˆˆ‰I•Í½±Ù”Ñ¡”¹•…É•ÍĞ½±‘•¸5½‘•°M™Í•Ñ¥½¸İ¥Ñ¡½ÕĞÉ½Ü¡…É‘½‘¥¹œ¸ˆˆˆ4(€€€€€€€™½È…¹‘¥‘…Ñ”¥¸É…¹”¡É½Ü°µ…à À°É½Ü€´€ÈÀÀ¤°€´Ä¤è4(€€€€€€€€€€€Ù…±Õ”€ôÍÑÈ¡Í•±˜¹É…İ}Ù…±Õ”¡˜‰	í…¹‘¥‘…Ñ•ôˆ¤½È€ˆˆ¤¹ÍÑÉ¥À ¤4(€€€€€€€€€€€¥˜Ù…±Õ”¥¸ì‹¶2C®“®æˆ°€‹²vó®ÂcªÒ®š³®æ‰ôè4(€€€€€€€€€€€€€€€É•ÑÕÉ¸Ù…±Õ”4(€€€€€€€É•ÑÕÉ¸€‹¶2CªÒ®æˆ4(4(€€€‘•˜}…Õ‘¥Ñ}¥Ñ•´¡Í•±˜°¥Ñ•´è¡…¹•1½œ¤€´øÍÑÈè4(€€€€€€€Í½ÕÉ”€ô¥Ñ•´¹Í½ÕÉ”4(€€€€€€€Á…ÉÑÌ€ôÍ½ÕÉ”¹ÍÁ±¥Ğ ˆ¸ˆ¤4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰Í…±•Ì¸ˆ¤…¹±•¸¡Á…ÉÑÌ¤€øô€Ìè4(€€€€€€€€€€€ÍÕ™™¥à€ôì4(€€€€€€€€€€€€€€€€‰ÅÕ…¹Ñ¥Ñäˆè€‹²"c®~$ˆ°4(€€€€€€€€€€€€€€€€‰…µ½Õ¹Ğˆè€‹®“²Ús²V„ˆ°4(€€€€€€€€€€€€€€€€‰µ…¹Õ™…ÑÕÉ•‘}ÅÕ…¹Ñ¥Ñäˆè€‹²‚s²†Àƒ²"c®~$ˆ°4(€€€€€€€€€€€€€€€€‰µ…¹Õ™…ÑÕÉ•‘}…µ½Õ¹Ğˆè€‹²‚s²†Àƒ®“²Ús²V„ˆ°4(€€€€€€€€€€€€€€€€‰½½‘Í}ÅÕ…¹Ñ¥Ñäˆè€‹²¶J ƒ²"c®~$ˆ°4(€€€€€€€€€€€€€€€€‰½½‘Í}…µ½Õ¹Ğˆè€‹²¶J ƒ®“²Ús²V„ˆ°4(€€€€€€€€€€€ô¹•Ğ¡Á…ÉÑÍl´Åt°Á…ÉÑÍl´Åt¤4(€€€€€€€€€€€ÁÉ½‘ÕĞ€ôì4(€€€€€€€€€€€€€€€€‰¹•İ}‰ÕÍ¥¹•ÍÌˆè€‹².ƒ²
³²^ˆ°4(€€€€€€€€€€€€€€€€‰½Ñ¡•Èˆè€‹ªâÃ¶®“²Úpˆ°4(€€€€€€€€€€€ô¹•Ğ¡Á…ÉÑÍlÅt°Á…ÉÑÍlÅt¤4(€€€€€€€€€€€É•ÑÕÉ¸˜‰íÁÉ½‘ÕÑôíÍÕ™™¥áôˆ4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰ÁÉ½‘ÕÑ¥½¸¸ˆ¤…¹±•¸¡Á…ÉÑÌ¤€øô€Èè4(€€€€€€€€€€€É•ÑÕÉ¸˜‰íÁ…ÉÑÍlÅuôƒ²w²
Ã²"c®~$ˆ4(€€€€€€€¥˜Í½ÕÉ”¹ÍÑ…ÉÑÍİ¥Ñ  ‰µ´¸ˆ¤…¹±•¸¡Á…ÉÑÌ¤€øô€Èè4(€€€€€€€€€€€É•ÑÕÉ¸˜‰íÁ…ÉÑÍlÅuôƒ²"c®~$ˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰É…İ}µ…Ñ•É¥…°¹ÁÕÉ¡…Í•}•ÍÑ¥µ…Ñ”¹™É½¹Ñ}ÁÉ½•ÍÌˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹ªÖ³®“¶2 ƒ²b#²ƒ¶"³²z®æ²‚ªÎ×²‚Tˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰É…İ}µ…Ñ•É¥…°¹ÁÕÉ¡…Í•}•ÍÑ¥µ…Ñ”¹‰…­}ÁÉ½•ÍÌˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹ªÖ³®“¶2 ƒ²b#²ƒ¶"³²z®æ¶nªÎ×²‚Tˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰½Ì¹‘¥ÍÁ½Í…°ˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹²‚s¶J ƒ¶>CªâÃ²C².ˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰½Ì¹½‰Í½±•Í•¹”ˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹²‚s¶J ƒ²®Ú¶fPƒ¶>'ªÂ²C².ˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰½Ì¹ÕÍÑ½µÍ}É•™Õ¹ˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹²nC²z³®0ƒªÒ²àƒ¶fcªâ'ªâ ˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰½Ì¹½½‘Ìˆè4(€€€€€€€€€€€É•ÑÕÉ¸€‹²¶J ƒ®“²Ús²nCªÂ ˆ4(€€€€€€€¥˜Í½ÕÉ”€ôô€‰Í…}…‘©ÕÍÑµ•¹Ğˆè4(€€€€€€€€€€€É½Ü€ô¥¹Ğ¡É”¹Í•…É ¡È‰q¬ˆ°¥Ñ•´¹•±°¤¹É½ÕÀ À¤¤4(€€€€€€€€€€€ÁÉ•™¥à€ôÍ•±˜¹}Í…}Í•Ñ¥½¸¡É½Ü¤4(€€€€€€€€€€€É•ÑÕÉ¸˜‰íÁÉ•™¥áõ}íÍ•±˜¹}É½İ}±…‰•°¡É½Ü¥ôˆ4(€€€€€€€É½Ü€ô¥¹Ğ¡É”¹Í•…É ¡È‰q¬ˆ°¥Ñ•´¹•±°¤¹É½ÕÀ À¤¤4(€€€€€€€É•ÑÕÉ¸Í•±˜¹}É½İ}±…‰•°¡É½Ü¤4(4(€€€ÍÑ…Ñ¥µ•Ñ¡½4(€€€‘•˜}…Õ‘¥Ñ}µ½¹Ñ ¡•±°èÍÑÈ¤€´ø¥¹Ğğ9½¹”è4(€€€€€€€µ…Ñ €ôÉ”¹™Õ±±µ…Ñ ¡Èˆ¡mµit¬¥q¬ˆ°•±°¤4(€€€€€€€¥˜¹½Ğµ…Ñ è4(€€€€€€€€€€€É•ÑÕÉ¸9½¹”4(€€€€€€€µ½¹Ñ €ô½±}¹Õµ‰•È¡µ…Ñ ¹É½ÕÀ Ä¤¤€´½±}¹Õµ‰•È ‰ˆ¤€¬€Ä4(€€€€€€€É•ÑÕÉ¸µ½¹Ñ ¥˜€Ä€ğôµ½¹Ñ €ğô€ÄÈ•±Í”9½¹”4(4(€€€ÍÑ…Ñ¥µ•Ñ¡½4(€€€‘•˜}…ÁÁ•¹‘}¥¹±¥¹•}•±°¡É½ÜèP¹±•µ•¹Ğ°…‘‘É•ÍÌèÍÑÈ°Ù…±Õ”è¹ä¤€´ø9½¹”è4(€€€€€€€¥˜¥Í¥¹ÍÑ…¹”¡Ù…±Õ”°€¡¥¹Ğ°™±½…Ğ¤¤…¹¹½Ğ¥Í¥¹ÍÑ…¹”¡Ù…±Õ”°‰½½°¤è4(€€€€€€€€€€€•±°€ôP¹MÕ‰±•µ•¹Ğ¡É½Ü°D ‰Œˆ¤°ì‰Èˆè…‘‘É•ÍÍô¤4(€€€€€€€€€€€P¹MÕ‰±•µ•¹Ğ¡•±°°D ‰Øˆ¤¤¹Ñ•áĞ€ôÉ•ÁÈ¡™±½…Ğ¡Ù…±Õ”¤¤4(€€€€€€€€€€€É•ÑÕÉ¸4(€€€€€€€•±°€ôP¹MÕ‰±•µ•¹Ğ¡É½Ü°D ‰Œˆ¤°ì‰Èˆè…‘‘É•ÍÌ°€‰Ğˆè€‰¥¹±¥¹•MÑÈ‰ô¤4(€€€€€€€¥¹±¥¹”€ôP¹MÕ‰±•µ•¹Ğ¡•±°°D ‰¥Ìˆ¤¤4(€€€€€€€Ñ•áĞ€ôP¹MÕ‰±•µ•¹Ğ¡¥¹±¥¹”°D ‰Ğˆ¤¤4(€€€€€€€Ñ•áĞ¹Ñ•áĞ€ô€ˆˆ¥˜Ù…±Õ”¥Ì9½¹”•±Í”ÍÑÈ¡Ù…±Õ”¤4(4(€€€ÍÑ…Ñ¥µ•Ñ¡½4(€€€‘•˜}İ½É­Í¡••Ñ}Ñ…É•Ğ¡•¹ÑÉ¥•Ìè‘¥ÑmÍÑÈ°‰åÑ•Ít°Í¡••Ñ}¹…µ”èÍÑÈ¤€´øÑÕÁ±•mÍÑÈ°P¹±•µ•¹Ğ°P¹±•µ•¹Ñtğ9½¹”è4(€€€€€€€İ½É­‰½½¬€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•Íl‰á°½İ½É­‰½½¬¹áµ°‰t¤4(€€€€€€€Í¡••ÑÌ€ôİ½É­‰½½¬¹™¥¹¡D ‰Í¡••ÑÌˆ¤¤4(€€€€€€€¥˜Í¡••ÑÌ¥Ì9½¹”è4(€€€€€€€€€€€É•ÑÕÉ¸9½¹”4(€€€€€€€Í¡••Ğ€ô¹•áĞ ¡¹½‘”™½È¹½‘”¥¸Í¡••ÑÌ¹™¥¹‘…±°¡D ‰Í¡••Ğˆ¤¤¥˜¹½‘”¹…ÑÑÉ¥ˆ¹•Ğ ‰¹…µ”ˆ¤€ôôÍ¡••Ñ}¹…µ”¤°9½¹”¤4(€€€€€€€¥˜Í¡••Ğ¥Ì9½¹”è4(€€€€€€€€€€€É•ÑÕÉ¸9½¹”4(€€€€€€€É•±…Ñ¥½¹Í¡¥Á}¥€ôÍ¡••Ğ¹…ÑÑÉ¥ˆ¹•Ğ¡˜‰ííí=%}I1}9Mõõõ¥ˆ¤4(€€€€€€€É•±…Ñ¥½¹Í¡¥ÁÌ€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•Íl‰á°½}É•±Ì½İ½É­‰½½¬¹áµ°¹É•±Ì‰t¤4(€€€€€€€É•±…Ñ¥½¹Í¡¥À€ô¹•áĞ 4(€€€€€€€€€€€€ 4(€€€€€€€€€€€€€€€¹½‘”™½È¹½‘”¥¸É•±…Ñ¥½¹Í¡¥ÁÌ¹™¥¹‘…±°¡˜‰íííI1}9MõõõI•±…Ñ¥½¹Í¡¥Àˆ¤4(€€€€€€€€€€€€€€€¥˜¹½‘”¹…ÑÑÉ¥ˆ¹•Ğ ‰%ˆ¤€ôôÉ•±…Ñ¥½¹Í¡¥Á}¥4(€€€€€€€€€€€€¤°4(€€€€€€€€€€€9½¹”°4(€€€€€€€€¤4(€€€€€€€¥˜É•±…Ñ¥½¹Í¡¥À¥Ì9½¹”è4(€€€€€€€€€€€É•ÑÕÉ¸9½¹”4(€€€€€€€Ñ…É•Ğ€ôÉ•±…Ñ¥½¹Í¡¥À¹…ÑÑÉ¥‰l‰Q…É•Ğ‰t¹É•Á±…” ‰qpˆ°€ˆ¼ˆ¤4(€€€€€€€¥˜Ñ…É•Ğ¹ÍÑ…ÉÑÍİ¥Ñ  ˆ¼ˆ¤è4(€€€€€€€€€€€Á…Ñ €ôÑ…É•Ğ¹±ÍÑÉ¥À ˆ¼ˆ¤4(€€€€€€€•±Í”è4(€€€€€€€€€€€Á…Ñ €ô˜‰á°½íÑ…É•Ñôˆ4(€€€€€€€É•ÑÕÉ¸Á…Ñ °İ½É­‰½½¬°É•±…Ñ¥½¹Í¡¥ÁÌ4(4(€€€ÍÑ…Ñ¥µ•Ñ¡½4(€€€‘•˜}¹•İ}…Õ‘¥Ñ}Í¡••Ğ¡•¹ÑÉ¥•Ìè‘¥ÑmÍÑÈ°‰åÑ•Ít¤€´øÑÕÁ±•mÍÑÈ°P¹±•µ•¹Ñtè4(€€€€€€€İ½É­‰½½­}¹…µ”€ô€‰á°½İ½É­‰½½¬¹áµ°ˆ4(€€€€€€€É•±Í}¹…µ”€ô€‰á°½}É•±Ì½İ½É­‰½½¬¹áµ°¹É•±Ìˆ4(€€€€€€€İ½É­‰½½¬€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•Ímİ½É­‰½½­}¹…µ•t¤4(€€€€€€€É•±…Ñ¥½¹Í¡¥ÁÌ€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•ÍmÉ•±Í}¹…µ•t¤4(€€€€€€€Í¡••ÑÌ€ôİ½É­‰½½¬¹™¥¹¡D ‰Í¡••ÑÌˆ¤¤4(€€€€€€€¥˜Í¡••ÑÌ¥Ì9½¹”è4(€€€€€€€€€€€Í¡••ÑÌ€ôP¹MÕ‰±•µ•¹Ğ¡İ½É­‰½½¬°D ‰Í¡••ÑÌˆ¤¤4(4(€€€€€€€•á¥ÍÑ¥¹}Ñ…É•ÑÌ€ôì4(€€€€€€€€€€€¹½‘”¹…ÑÑÉ¥ˆ¹•Ğ ‰Q…É•Ğˆ°€ˆˆ¤4(€€€€€€€€€€€™½È¹½‘”¥¸É•±…Ñ¥½¹Í¡¥ÁÌ¹™¥¹‘…±°¡˜‰íííI1}9MõõõI•±…Ñ¥½¹Í¡¥Àˆ¤4(€€€€€€€ô4(€€€€€€€Í¡••Ñ}¹Õµ‰•È€ô€Ä4(€€€€€€€İ¡¥±”˜‰İ½É­Í¡••ÑÌ½Í¡••ÑíÍ¡••Ñ}¹Õµ‰•Éô¹áµ°ˆ¥¸•á¥ÍÑ¥¹}Ñ…É•ÑÌè4(€€€€€€€€€€€Í¡••Ñ}¹Õµ‰•È€¬ô€Ä4(€€€€€€€Ñ…É•Ğ€ô˜‰İ½É­Í¡••ÑÌ½Í¡••ÑíÍ¡••Ñ}¹Õµ‰•Éô¹áµ°ˆ4(€€€€€€€Í¡••Ñ}Á…Ñ €ô˜‰á°½íÑ…É•Ñôˆ4(4(€€€€€€€•á¥ÍÑ¥¹}É¥‘Ì€ôì4(€€€€€€€€€€€¹½‘”¹…ÑÑÉ¥ˆ¹•Ğ ‰%ˆ°€ˆˆ¤4(€€€€€€€€€€€™½È¹½‘”¥¸É•±…Ñ¥½¹Í¡¥ÁÌ¹™¥¹‘…±°¡˜‰íííI1}9MõõõI•±…Ñ¥½¹Í¡¥Àˆ¤4(€€€€€€€ô4(€€€€€€€É¥‘}¹Õµ‰•È€ô€Ä4(€€€€€€€İ¡¥±”˜‰É%‘íÉ¥‘}¹Õµ‰•Éôˆ¥¸•á¥ÍÑ¥¹}É¥‘Ìè4(€€€€€€€€€€€É¥‘}¹Õµ‰•È€¬ô€Ä4(€€€€€€€É•±…Ñ¥½¹Í¡¥Á}¥€ô˜‰É%‘íÉ¥‘}¹Õµ‰•Éôˆ4(€€€€€€€P¹MÕ‰±•µ•¹Ğ 4(€€€€€€€€€€€É•±…Ñ¥½¹Í¡¥ÁÌ°4(€€€€€€€€€€€˜‰íííI1}9MõõõI•±…Ñ¥½¹Í¡¥Àˆ°4(€€€€€€€€€€€ì4(€€€€€€€€€€€€€€€€‰%ˆèÉ•±…Ñ¥½¹Í¡¥Á}¥°4(€€€€€€€€€€€€€€€€‰QåÁ”ˆè˜‰í=%}I1}9Mô½İ½É­Í¡••Ğˆ°4(€€€€€€€€€€€€€€€€‰Q…É•ĞˆèÑ…É•Ğ°4(€€€€€€€€€€€ô°4(€€€€€€€€¤4(4(€€€€€€€Í¡••Ñ}¥‘Ì€ôm¥¹Ğ¡¹½‘”¹…ÑÑÉ¥ˆ¹•Ğ ‰Í¡••Ñ%ˆ°€ˆÀˆ¤¤™½È¹½‘”¥¸Í¡••ÑÌ¹™¥¹‘…±°¡D ‰Í¡••Ğˆ¤¥t4(€€€€€€€¹•İ}Í¡••Ğ€ôP¹±•µ•¹Ğ 4(€€€€€€€€€€€D ‰Í¡••Ğˆ¤°4(€€€€€€€€€€€ì4(€€€€€€€€€€€€€€€€‰¹…µ”ˆèU%Q}M!Q}95°4(€€€€€€€€€€€€€€€€‰Í¡••Ñ%ˆèÍÑÈ¡µ…à¡Í¡••Ñ}¥‘Ì°‘•™…Õ±ĞôÀ¤€¬€Ä¤°4(€€€€€€€€€€€€€€€˜‰ííí=%}I1}9Mõõõ¥ˆèÉ•±…Ñ¥½¹Í¡¥Á}¥°4(€€€€€€€€€€€ô°4(€€€€€€€€¤4(€€€€€€€¡¥±‘É•¸€ô±¥ÍĞ¡Í¡••ÑÌ¤4(€€€€€€€‘…Ñ…}¥¹‘•à€ô¹•áĞ ¡¥¹‘•à™½È¥¹‘•à°¹½‘”¥¸•¹Õµ•É…Ñ”¡¡¥±‘É•¸¤¥˜¹½‘”¹…ÑÑÉ¥ˆ¹•Ğ ‰¹…µ”ˆ¤€ôô€‰…Ñ„ˆ¤°±•¸¡¡¥±‘É•¸¤€´€Ä¤4(€€€€€€€Í¡••ÑÌ¹¥¹Í•ÉĞ¡µ…à¡‘…Ñ…}¥¹‘•à€¬€Ä°€À¤°¹•İ}Í¡••Ğ¤4(4(€€€€€€€½¹Ñ•¹Ñ}ÑåÁ•Í}¹…µ”€ô€‰m½¹Ñ•¹Ñ}QåÁ•Ít¹áµ°ˆ4(€€€€€€€½¹Ñ•¹Ñ}ÑåÁ•Ì€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•Ím½¹Ñ•¹Ñ}ÑåÁ•Í}¹…µ•t¤4(€€€€€€€P¹MÕ‰±•µ•¹Ğ 4(€€€€€€€€€€€½¹Ñ•¹Ñ}ÑåÁ•Ì°4(€€€€€€€€€€€˜‰ííí=9Q9Q}QeAM}9Mõõõ=Ù•ÉÉ¥‘”ˆ°4(€€€€€€€€€€€ì4(€€€€€€€€€€€€€€€€‰A…ÉÑ9…µ”ˆè˜ˆ½íÍ¡••Ñ}Á…Ñ¡ôˆ°4(€€€€€€€€€€€€€€€€‰½¹Ñ•¹ÑQåÁ”ˆè€‰…ÁÁ±¥…Ñ¥½¸½Ù¹¹½Á•¹áµ±™½Éµ…ÑÌµ½™™¥•‘½Õµ•¹Ğ¹ÍÁÉ•…‘Í¡••Ñµ°¹İ½É­Í¡••Ğ­áµ°ˆ°4(€€€€€€€€€€€ô°4(€€€€€€€€¤4(€€€€€€€•¹ÑÉ¥•Ímİ½É­‰½½­}¹…µ•t€ôÍ•É¥…±¥é•}áµ°¡İ½É­‰½½¬°€ ‰àÄÔˆ°€‰áÈˆ°€‰áÈØˆ°€‰áÈÄÀˆ°€‰áÈÈˆ¤¤4(€€€€€€€•¹ÑÉ¥•ÍmÉ•±Í}¹…µ•t€ôÍ•É¥…±¥é•}áµ°¡É•±…Ñ¥½¹Í¡¥ÁÌ°‘•™…Õ±Ñ}¹…µ•ÍÁ…”õI1}9L¤4(€€€€€€€•¹ÑÉ¥•Ím½¹Ñ•¹Ñ}ÑåÁ•Í}¹…µ•t€ôÍ•É¥…±¥é•}áµ°¡½¹Ñ•¹Ñ}ÑåÁ•Ì°‘•™…Õ±Ñ}¹…µ•ÍÁ…”õ=9Q9Q}QeAM}9L¤4(4(€€€€€€€É½½Ğ€ôP¹±•µ•¹Ğ¡D ‰İ½É­Í¡••Ğˆ¤¤4(€€€€€€€Ù¥•İÌ€ôP¹MÕ‰±•µ•¹Ğ¡É½½Ğ°D ‰Í¡••ÑY¥•İÌˆ¤¤4(€€€€€€€Ù¥•Ü€ôP¹MÕ‰±•µ•¹Ğ¡Ù¥•İÌ°D ‰Í¡••ÑY¥•Üˆ¤°ì‰İ½É­‰½½­Y¥•İ%ˆè€ˆÀ‰ô¤4(€€€€€€€P¹MÕ‰±•µ•¹Ğ¡Ù¥•Ü°D ‰Á…¹”ˆ¤°ì‰åMÁ±¥Ğˆè€ˆÄˆ°€‰Ñ½Á1•™Ñ•±°ˆè€‰Èˆ°€‰…Ñ¥Ù•A…¹”ˆè€‰‰½ÑÑ½µ1•™Ğˆ°€‰ÍÑ…Ñ”ˆè€‰™É½é•¸‰ô¤4(€€€€€€€P¹MÕ‰±•µ•¹Ğ¡É½½Ğ°D ‰Í¡••Ñ½Éµ…ÑAÈˆ¤°ì‰‘•™…Õ±ÑI½İ!•¥¡Ğˆè€ˆÄÔ‰ô¤4(€€€€€€€½±Õµ¹Ì€ôP¹MÕ‰±•µ•¹Ğ¡É½½Ğ°D ‰½±Ìˆ¤¤4(€€€€€€€™½È¥¹‘•à°İ¥‘Ñ ¥¸•¹Õµ•É…Ñ”  Ü°€ä°€Äà°€Èà°€ÄÌ°€ÄÜ°€ÄÜ°€ÄÜ°€ØÈ°€ÄÀ¤°ÍÑ…ÉĞôÄ¤è4(€€€€€€€€€€€P¹MÕ‰±•µ•¹Ğ¡½±Õµ¹Ì°D ‰½°ˆ¤°ì‰µ¥¸ˆèÍÑÈ¡¥¹‘•à¤°€‰µ…àˆèÍÑÈ¡¥¹‘•à¤°€‰İ¥‘Ñ ˆèÍÑÈ¡İ¥‘Ñ ¤°€‰ÕÍÑ½µ]¥‘Ñ ˆè€ˆÄ‰ô¤4(€€€€€€€Í¡••Ñ}‘…Ñ„€ôP¹MÕ‰±•µ•¹Ğ¡É½½Ğ°D ‰Í¡••Ñ…Ñ„ˆ¤¤4(€€€€€€€¡•…‘•È€ôP¹MÕ‰±•µ•¹Ğ¡Í¡••Ñ}‘…Ñ„°D ‰É½Üˆ¤°ì‰Èˆè€ˆÄ‰ô¤4(€€€€€€€™½È¥¹‘•à°Ù…±Õ”¥¸•¹Õµ•É…Ñ”  ‹®Ê#¶bàˆ°€‹²nPˆ°€‹ªÖ³®Úˆ°€‹¶V·®ª¤ˆ°€‹®Âc²bƒ² ˆ°€‹ªâÃ²’ªÂHˆ°€‹²z®‚—ªÂHˆ°€‹²Â£²vĞˆ°€‹²
³²r€ˆ°€‹² ƒ²vÓ®>dˆ¤°ÍÑ…ÉĞôÄ¤è4(€€€€€€€€€€€½±‘•¹]½É­‰½½¬¹}…ÁÁ•¹‘}¥¹±¥¹•}•±°¡¡•…‘•È°˜‰í½±}¹…µ”¡¥¹‘•à¥ôÄˆ°Ù…±Õ”¤4(€€€€€€€P¹MÕ‰±•µ•¹Ğ¡É½½Ğ°D ‰…ÕÑ½¥±Ñ•Èˆ¤°ì‰É•˜ˆè€‰Äé(Ä‰ô¤4(€€€€€€€P¹MÕ‰±•µ•¹Ğ¡É½½Ğ°D ‰¡åÁ•É±¥¹­Ìˆ¤¤4(€€€€€€€É•ÑÕÉ¸Í¡••Ñ}Á…Ñ °É½½Ğ4(4(€€€‘•˜}…ÁÁ•¹‘}…Õ‘¥Ñ}Í¡••Ğ¡Í•±˜°•¹ÑÉ¥•Ìè‘¥ÑmÍÑÈ°‰åÑ•Ít¤€´ø9½¹”è4(€€€€€€€‘¥É•Ñ}±½Ì€ôÍ•±˜¹}‘¥É•Ñ}¥¹ÁÕÑ}±½Ì ¤4(€€€€€€€•á¥ÍÑ¥¹œ€ôÍ•±˜¹}İ½É­Í¡••Ñ}Ñ…É•Ğ¡•¹ÑÉ¥•Ì°U%Q}M!Q}95¤4(€€€€€€€¥˜•á¥ÍÑ¥¹œ¥Ì9½¹”è4(€€€€€€€€€€€Í¡••Ñ}Á…Ñ °…Õ‘¥Ñ}É½½Ğ€ôÍ•±˜¹}¹•İ}…Õ‘¥Ñ}Í¡••Ğ¡•¹ÑÉ¥•Ì¤4(€€€€€€€•±Í”è4(€€€€€€€€€€€Í¡••Ñ}Á…Ñ €ô•á¥ÍÑ¥¹lÁt4(€€€€€€€€€€€…Õ‘¥Ñ}É½½Ğ€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•ÍmÍ¡••Ñ}Á…Ñ¡t¤4(€€€€€€€¥˜¹½Ğ‘¥É•Ñ}±½Ìè4(€€€€€€€€€€€•¹ÑÉ¥•ÍmÍ¡••Ñ}Á…Ñ¡t€ôÍ•É¥…±¥é•}áµ°¡…Õ‘¥Ñ}É½½Ğ¤4(€€€€€€€€€€€É•ÑÕÉ¸4(4(€€€€€€€Í¡••Ñ}‘…Ñ„€ô…Õ‘¥Ñ}É½½Ğ¹™¥¹¡D ‰Í¡••Ñ…Ñ„ˆ¤¤4(€€€€€€€¥˜Í¡••Ñ}‘…Ñ„¥Ì9½¹”è4(€€€€€€€€€€€Í¡••Ñ}‘…Ñ„€ôP¹MÕ‰±•µ•¹Ğ¡…Õ‘¥Ñ}É½½Ğ°D ‰Í¡••Ñ…Ñ„ˆ¤¤4(€€€€€€€•á¥ÍÑ¥¹}É½İÌ€ôm¥¹Ğ¡É½Ü¹…ÑÑÉ¥ˆ¹•Ğ ‰Èˆ°€ˆÀˆ¤¤™½ÈÉ½Ü¥¸Í¡••Ñ}‘…Ñ„¹™¥¹‘…±°¡D ‰É½Üˆ¤¥t4(€€€€€€€¹•áÑ}É½Ü€ôµ…à¡•á¥ÍÑ¥¹}É½İÌ°‘•™…Õ±ĞôÄ¤€¬€Ä4(€€€€€€€¡åÁ•É±¥¹­Ì€ô…Õ‘¥Ñ}É½½Ğ¹™¥¹¡D ‰¡åÁ•É±¥¹­Ìˆ¤¤4(€€€€€€€¥˜¡åÁ•É±¥¹­Ì¥Ì9½¹”è4(€€€€€€€€€€€¡åÁ•É±¥¹­Ì€ôP¹MÕ‰±•µ•¹Ğ¡…Õ‘¥Ñ}É½½Ğ°D ‰¡åÁ•É±¥¹­Ìˆ¤¤4(4(€€€€€€€™½È¥Ñ•´¥¸‘¥É•Ñ}±½Ìè4(€€€€€€€€€€€É½Ü€ôP¹MÕ‰±•µ•¹Ğ¡Í¡••Ñ}‘…Ñ„°D ‰É½Üˆ¤°ì‰ÈˆèÍÑÈ¡¹•áÑ}É½Ü¥ô¤4(€€€€€€€€€€€µ½¹Ñ €ôÍ•±˜¹}…Õ‘¥Ñ}µ½¹Ñ ¡¥Ñ•´¹•±°¤4(€€€€€€€€€€€‘•±Ñ„€ô€ 4(€€€€€€€€€€€€€€€™±½…Ğ¡¥Ñ•´¹¹•İ}Ù…±Õ”¤€´™±½…Ğ¡¥Ñ•´¹½±‘}Ù…±Õ”¤4(€€€€€€€€€€€€€€€¥˜¥Í¥¹ÍÑ…¹”¡¥Ñ•´¹½±‘}Ù…±Õ”°€¡¥¹Ğ°™±½…Ğ¤¤…¹¥Í¥¹ÍÑ…¹”¡¥Ñ•´¹¹•İ}Ù…±Õ”°€¡¥¹Ğ°™±½…Ğ¤¤4(€€€€€€€€€€€€€€€•±Í”€ˆˆ4(€€€€€€€€€€€€¤4(€€€€€€€€€€€Ù…±Õ•Ì€ô€ 4(€€€€€€€€€€€€€€€¹•áÑ}É½Ü€´€Ä°4(€€€€€€€€€€€€€€€˜‰íµ½¹Ñ¡÷²nPˆ¥˜µ½¹Ñ •±Í”€ˆˆ°4(€€€€€€€€€€€€€€€Í•±˜¹}…Õ‘¥Ñ}…Ñ•½Éä¡¥Ñ•´¹Í½ÕÉ”¤°4(€€€€€€€€€€€€€€€Í•±˜¹}…Õ‘¥Ñ}¥Ñ•´¡¥Ñ•´¤°4(€€€€€€€€€€€€€€€¥Ñ•´¹•±°°4(€€€€€€€€€€€€€€€¥Ñ•´¹½±‘}Ù…±Õ”°4(€€€€€€€€€€€€€€€¥Ñ•´¹¹•İ}Ù…±Õ”°4(€€€€€€€€€€€€€€€‘•±Ñ„°4(€€€€€€€€€€€€€€€¥Ñ•´¹É•…Í½¸°4(€€€€€€€€€€€€€€€€‹²vÓ®>dˆ°4(€€€€€€€€€€€€¤4(€€€€€€€€€€€™½È¥¹‘•à°Ù…±Õ”¥¸•¹Õµ•É…Ñ”¡Ù…±Õ•Ì°ÍÑ…ÉĞôÄ¤è4(€€€€€€€€€€€€€€€Í•±˜¹}…ÁÁ•¹‘}¥¹±¥¹•}•±°¡É½Ü°˜‰í½±}¹…µ”¡¥¹‘•à¥õí¹•áÑ}É½İôˆ°Ù…±Õ”¤4(€€€€€€€€€€€P¹MÕ‰±•µ•¹Ğ 4(€€€€€€€€€€€€€€€¡åÁ•É±¥¹­Ì°4(€€€€€€€€€€€€€€€D ‰¡åÁ•É±¥¹¬ˆ¤°4(€€€€€€€€€€€€€€€ì4(€€€€€€€€€€€€€€€€€€€€‰É•˜ˆè˜‰)í¹•áÑ}É½İôˆ°4(€€€€€€€€€€€€€€€€€€€€‰±½…Ñ¥½¸ˆè˜ˆ…Ñ„œ…í¥Ñ•´¹•±±ôˆ°4(€€€€€€€€€€€€€€€€€€€€‰‘¥ÍÁ±…äˆè€‹²vÓ®>dˆ°4(€€€€€€€€€€€€€€€ô°4(€€€€€€€€€€€€¤4(€€€€€€€€€€€¹•áÑ}É½Ü€¬ô€Ä4(4(€€€€€€€…ÕÑ½}™¥±Ñ•È€ô…Õ‘¥Ñ}É½½Ğ¹™¥¹¡D ‰…ÕÑ½¥±Ñ•Èˆ¤¤4(€€€€€€€¥˜…ÕÑ½}™¥±Ñ•È¥Ì9½¹”è4(€€€€€€€€€€€…ÕÑ½}™¥±Ñ•È€ôP¹MÕ‰±•µ•¹Ğ¡…Õ‘¥Ñ}É½½Ğ°D ‰…ÕÑ½¥±Ñ•Èˆ¤¤4(€€€€€€€…ÕÑ½}™¥±Ñ•È¹…ÑÑÉ¥‰l‰É•˜‰t€ô˜‰Äé)í¹•áÑ}É½Ü€´€Åôˆ4(€€€€€€€•¹ÑÉ¥•ÍmÍ¡••Ñ}Á…Ñ¡t€ôÍ•É¥…±¥é•}áµ°¡…Õ‘¥Ñ}É½½Ğ¤4(4(€€€‘•˜Í…Ù”¡Í•±˜°‘•ÍÑ¥¹…Ñ¥½¸èÍÑÈğA…Ñ ¤€´øA…Ñ è4(€€€€€€€‘•ÍÑ¥¹…Ñ¥½¸€ôA…Ñ ¡‘•ÍÑ¥¹…Ñ¥½¸¤4(€€€€€€€‘•ÍÑ¥¹…Ñ¥½¸¹Á…É•¹Ğ¹µ­‘¥È¡Á…É•¹ÑÌõQÉÕ”°•á¥ÍÑ}½¬õQÉÕ”¤4(€€€€€€€Í•±˜¹É•…±Õ±…Ñ” ¤4(€€€€€€€€Œ%˜…¸¥¹ÁÕĞÉ•Á±…•ÌÑ¡”…¹¡½È½˜„Í¡…É•™½ÉµÕ±„°‘•Ñ… ½¹±äÑ¡…Ğ4(€€€€€€€€ŒÍ¡…É•É½ÕÀ¸=Ñ¡•È™½ÉµÕ±…Ì…¹…¡•Ù…±Õ•ÌÉ•µ…¥¸‰åÑ”µ•ÅÕ¥Ù…±•¹Ğ4(€€€€€€€€Œ¥¸µ•…¹¥¹œÑ¼Ñ¡”½±‘•¸5½‘•°°İ¡¥ ¥ÌÑ¡”µ½ÍĞá•°µ½µÁ…Ñ¥‰±”4(€€€€€€€€Œİ…äÑ¼ÁÉ•Í•ÉÙ”„½µÁ±•àİ½É­‰½½¬¸4(€€€€€€€¥µÁ…Ñ•‘}Í¡…É•èÍ•ÑmÍÑÉt€ôÍ•Ğ ¤4(€€€€€€€™½È…‘‘È¥¸Í•±˜¹½Ù•ÉÉ¥‘•Ìè4(€€€€€€€€€€€•±°€ôÍ•±˜¹•±±Ì¹•Ğ¡…‘‘È¤4(€€€€€€€€€€€˜€ô•±°¹™¥¹¡D ‰˜ˆ¤¤¥˜•±°¥Ì¹½Ğ9½¹”•±Í”9½¹”4(€€€€€€€€€€€¥˜˜¥Ì¹½Ğ9½¹”…¹˜¹…ÑÑÉ¥ˆ¹•Ğ ‰Ğˆ¤€ôô€‰Í¡…É•ˆ…¹˜¹…ÑÑÉ¥ˆ¹•Ğ ‰Í¤ˆ¤…¹€¡˜¹Ñ•áĞ½È€ˆˆ¤¹ÍÑÉ¥À ¤è4(€€€€€€€€€€€€€€€¥µÁ…Ñ•‘}Í¡…É•¹…‘¡˜¹…ÑÑÉ¥‰l‰Í¤‰t¤4(4(€€€€€€€™½È…‘‘È°•±°¥¸Í•±˜¹•±±Ì¹¥Ñ•µÌ ¤è4(€€€€€€€€€€€˜€ô•±°¹™¥¹¡D ‰˜ˆ¤¤4(€€€€€€€€€€€¥˜˜¥Ì9½¹”½È˜¹…ÑÑÉ¥ˆ¹•Ğ ‰Í¤ˆ¤¹½Ğ¥¸¥µÁ…Ñ•‘}Í¡…É•½È…‘‘È¥¸Í•±˜¹½Ù•ÉÉ¥‘•Ìè4(€€€€€€€€€€€€€€€½¹Ñ¥¹Õ”4(€€€€€€€€€€€˜¹…ÑÑÉ¥ˆ¹±•…È ¤4(€€€€€€€€€€€˜¹Ñ•áĞ€ôÍ•±˜¹™½ÉµÕ±…Ím…‘‘Ét¹±ÍÑÉ¥À ˆôˆ¤4(4(€€€€€€€™½È…‘‘È°Ù…±Õ”¥¸Í•±˜¹½Ù•ÉÉ¥‘•Ì¹¥Ñ•µÌ ¤è4(€€€€€€€€€€€•±°€ôÍ•±˜¹•±±Ì¹•Ğ¡…‘‘È¤4(€€€€€€€€€€€¥˜•±°¥Ì9½¹”è4(€€€€€€€€€€€€€€€É½İ}¹Õ´€ôÉ”¹Í•…É ¡È‰q¬ˆ°…‘‘È¤¹É½ÕÀ À¤4(€€€€€€€€€€€€€€€Í¡••Ñ}‘…Ñ„€ôÍ•±˜¹É½½Ğ¹™¥¹¡D ‰Í¡••Ñ…Ñ„ˆ¤¤4(€€€€€€€€€€€€€€€É½Ü€ô¹•áĞ ¡È™½ÈÈ¥¸Í¡••Ñ}‘…Ñ„¹™¥¹‘…±°¡D ‰É½Üˆ¤¤¥˜È¹…ÑÑÉ¥ˆ¹•Ğ ‰Èˆ¤€ôôÉ½İ}¹Õ´¤°9½¹”¤4(€€€€€€€€€€€€€€€¥˜É½Ü¥Ì9½¹”è4(€€€€€€€€€€€€€€€€€€€É½Ü€ôP¹MÕ‰±•µ•¹Ğ¡Í¡••Ñ}‘…Ñ„°D ‰É½Üˆ¤°ì‰ÈˆèÉ½İ}¹Õµô¤4(€€€€€€€€€€€€€€€•±°€ôP¹MÕ‰±•µ•¹Ğ¡É½Ü°D ‰Œˆ¤°ì‰Èˆè…‘‘Éô¤4(€€€€€€€€€€€€€€€Í•±˜¹•±±Ím…‘‘Ét€ô•±°4(€€€€€€€€€€€™½È¡¥±¥¸±¥ÍĞ¡•±°¤è4(€€€€€€€€€€€€€€€¥˜¡¥±¹Ñ…œ¥¸€¡D ‰˜ˆ¤°D ‰Øˆ¤°D ‰¥Ìˆ¤¤è•±°¹É•µ½Ù”¡¡¥±¤4(€€€€€€€€€€€¥˜¥Í¥¹ÍÑ…¹”¡Ù…±Õ”°ÍÑÈ¤è4(€€€€€€€€€€€€€€€•±°¹…ÑÑÉ¥‰l‰Ğ‰t€ô€‰¥¹±¥¹•MÑÈˆ4(€€€€€€€€€€€€€€€¥¹±¥¹”€ôP¹MÕ‰±•µ•¹Ğ¡•±°°D ‰¥Ìˆ¤¤4(€€€€€€€€€€€€€€€P¹MÕ‰±•µ•¹Ğ¡¥¹±¥¹”°D ‰Ğˆ¤¤¹Ñ•áĞ€ôÙ…±Õ”4(€€€€€€€€€€€•±Í”è4(€€€€€€€€€€€€€€€•±°¹…ÑÑÉ¥ˆ¹Á½À ‰Ğˆ°9½¹”¤4(€€€€€€€€€€€€€€€P¹MÕ‰±•µ•¹Ğ¡•±°°D ‰Øˆ¤¤¹Ñ•áĞ€ôÉ•ÁÈ¡™±½…Ğ¡Ù…±Õ”¤¤4(€€€€€€€•¹ÑÉ¥•Ì€ô‘¥Ğ¡Í•±˜¹•¹ÑÉ¥•Ì¤4(€€€€€€€•¹ÑÉ¥•ÍmÍ•±˜¹Í¡••Ñ}áµ±t€ôÍ•É¥…±¥é•}áµ°¡Í•±˜¹É½½Ğ°€ ‰àÄÑ…Œˆ°€‰áÈˆ°€‰áÈÈˆ°€‰áÈÌˆ¤¤4(€€€€€€€Í•±˜¹}…ÁÁ•¹‘}…Õ‘¥Ñ}Í¡••Ğ¡•¹ÑÉ¥•Ì¤4(€€€€€€€¥˜Í•±˜¹}İ½É­Í¡••Ñ}Ñ…É•Ğ¡•¹ÑÉ¥•Ì°U%Q}M!Q}95¤¥Ì9½¹”è4(€€€€€€€€€€€É…¥Í”IÕ¹Ñ¥µ•ÉÉ½È ‹²z®‚—®Âc²b®
Ó²^´ƒ².s¶*ã®–ğƒ²w²Ç¶Vc² ƒ®ªï¶Z#²*×®.#®.¸ˆ¤4(€€€€€€€İ½É­‰½½­}¹…µ”€ô€‰á°½İ½É­‰½½¬¹áµ°ˆ4(€€€€€€€İˆ€ôP¹™É½µÍÑÉ¥¹œ¡•¹ÑÉ¥•Ímİ½É­‰½½­}¹…µ•t¤4(€€€€€€€…±Œ€ôİˆ¹™¥¹¡D ‰…±AÈˆ¤¤4(€€€€€€€¥˜…±Œ¥Ì9½¹”è…±Œ€ôP¹MÕ‰±•µ•¹Ğ¡İˆ°D ‰…±AÈˆ¤¤4(€€€€€€€…±Œ¹…ÑÑÉ¥ˆ¹ÕÁ‘…Ñ”¡ì‰…±5½‘”ˆè€‰…ÕÑ¼ˆ°€‰™Õ±±…±=¹1½…ˆè€ˆÄˆ°€‰™½É•Õ±±…±Œˆè€ˆÄ‰ô¤4(€€€€€€€•¹ÑÉ¥•Ímİ½É­‰½½­}¹…µ•t€ôÍ•É¥…±¥é•}áµ°¡İˆ°€ ‰àÄÔˆ°€‰áÈˆ°€‰áÈØˆ°€‰áÈÄÀˆ°€‰áÈÈˆ¤¤4(€€€€€€€İ¥Ñ é¥Á™¥±”¹i¥Á¥±”¡‘•ÍÑ¥¹…Ñ¥½¸°€‰Üˆ°é¥Á™¥±”¹i%A}1Q¤…Ì…É¡¥Ù”è4(€€€€€€€€€€€™½È¹…µ”°‘…Ñ„¥¸•¹ÑÉ¥•Ì¹¥Ñ•µÌ ¤è…É¡¥Ù”¹İÉ¥Ñ•ÍÑÈ¡¹…µ”°‘…Ñ„¤4(€€€€€€€É•ÑÕÉ¸‘•ÍÑ¥¹…Ñ¥½¸4(4(€€€‘•˜±½}‘¥ÑÌ¡Í•±˜¤€´ø±¥ÍÑm‘¥ÑmÍÑÈ°¹åutè4(€€€€€€€É•ÑÕÉ¸m…Í‘¥Ğ¡¥Ñ•´¤™½È¥Ñ•´¥¸Í•±˜¹±½t4(