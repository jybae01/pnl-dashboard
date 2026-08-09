from __future__ import annotations

import re
import zipfile
from numbers import Real
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping

from .workbook import extract_period_types, infer_workbook_year


_SPACE = re.compile(r"\s+")
_ALLOWED_PERIOD_TYPES = {"실적", "추정", "계획"}
_MONTH_HEADER = re.compile(r"(?<!\d)(\d{1,2})\s*월")


def _normalized(value: Any) -> str:
    return _SPACE.sub("", str(value or "")).casefold()


@dataclass(frozen=True)
class AnchorSpec:
    code: str
    labels: tuple[str, ...]
    expected_row: int
    tolerance: int = 0
    columns: tuple[str, ...] = ("A", "B", "C", "D")


@dataclass(frozen=True)
class AnchorBlockSpec:
    code: str
    start: AnchorSpec
    stop: AnchorSpec
    mapped_rows: tuple[int, ...]


@dataclass(frozen=True)
class NumericRowSpec:
    code: str
    rows: tuple[int, ...]
    allow_blank: bool = False


@dataclass(frozen=True)
class PreflightIssue:
    code: str
    message: str
    expected: Any = None
    observed: Any = None
    location: str | None = None


@dataclass
class PreflightReport:
    path: str
    passed: bool
    issues: list[PreflightIssue] = field(default_factory=list)
    anchors: dict[str, int] = field(default_factory=dict)
    workbook_year: int | None = None
    period_types: dict[str, str] = field(default_factory=dict)

    def as_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "passed": self.passed,
            "issues": [asdict(issue) for issue in self.issues],
            "anchors": dict(self.anchors),
            "workbook_year": self.workbook_year,
            "period_types": dict(self.period_types),
        }


class PreflightValidationError(ValueError):
    def __init__(self, report: PreflightReport):
        self.report = report
        summary = "; ".join(issue.message for issue in report.issues[:5])
        super().__init__(summary or "workbook pre-flight validation failed")


class ExcelPreflightValidator:
    """Mapping-aware structural gate executed before deterministic calculation.

    The gate verifies anchor identity, relative block ordering and that every
    mapped source row remains inside its approved block. It never changes the
    workbook and it does not calculate accounting values.
    """

    MAX_WORKBOOK_BYTES = 50 * 1024 * 1024

    def __init__(
        self,
        mapping: Mapping[str, Any],
        *,
        anchors: Iterable[AnchorSpec] | None = None,
        blocks: Iterable[AnchorBlockSpec] | None = None,
        numeric_rows: Iterable[NumericRowSpec] | None = None,
    ):
        self.mapping = dict(mapping)
        discovery = self.mapping.get("analysis_adapter", {}).get("account_discovery", {})
        default_anchors = (
            AnchorSpec(
                "front_raw_material_total",
                ("전공정 원재료 합계", "전공정원재료합계", "생산출고 계"),
                int(self.mapping.get("analysis_adapter", {}).get("material", {})
                    .get("front_process", {}).get("total_amount_row", 211)),
                tolerance=0,
            ),
            AnchorSpec(
                "back_raw_material_total",
                ("후공정 원재료 합계", "후공정원재료합계", "생산출고 계"),
                int(self.mapping.get("analysis_adapter", {}).get("material", {})
                    .get("back_process", {}).get("total_amount_row", 699)),
                tolerance=0,
            ),
            AnchorSpec(
                "manufacturing_start",
                (str(discovery.get("manufacturing_start_marker") or "★제조경비 명세서_입력"),),
                287,
                tolerance=0,
            ),
            AnchorSpec(
                "manufacturing_stop",
                (str(discovery.get("manufacturing_stop_marker") or "*제조원가 변동비/고정비 비율"),),
                321,
                tolerance=0,
            ),
            AnchorSpec(
                "sga_start",
                (str(discovery.get("sga_start_marker") or "★판매관리비 관리 명세서"),),
                1166,
                tolerance=0,
            ),
            AnchorSpec(
                "sga_stop",
                (str(discovery.get("sga_stop_marker") or "★손익계산서"),),
                1244,
                tolerance=0,
            ),
            AnchorSpec("operating_profit", ("영업이익",), 1306, tolerance=0),
        )
        self.anchors = tuple(anchors or default_anchors)
        by_code = {item.code: item for item in self.anchors}
        default_blocks: tuple[AnchorBlockSpec, ...] = ()
        if {"manufacturing_start", "manufacturing_stop"}.issubset(by_code):
            default_blocks += (AnchorBlockSpec(
                "manufacturing_accounts",
                by_code["manufacturing_start"],
                by_code["manufacturing_stop"],
                tuple(int(row) for row in self.mapping.get("manufacturing_input_rows", ())),
            ),)
        if {"sga_start", "sga_stop"}.issubset(by_code):
            default_blocks += (AnchorBlockSpec(
                "sga_accounts",
                by_code["sga_start"],
                by_code["sga_stop"],
                tuple(int(row) for row in self.mapping.get("sga_input_rows", ())),
            ),)
        self.blocks = tuple(blocks or default_blocks)
        adapter = self.mapping.get("analysis_adapter", {})
        material = adapter.get("material", {})
        manufacturing = adapter.get("manufacturing", {})
        comparison = self.mapping.get("comparison", {})
        sales_groups = comparison.get("sales_groups", {})
        pnl_rows = comparison.get("pnl_rows", {})
        default_numeric_rows = (
            NumericRowSpec("jpy_fx", (int(material.get("jpy_fx_row", 9)),)),
            NumericRowSpec("front_raw_material", tuple(range(205, 212))),
            NumericRowSpec("manufacturing_accounts", tuple(
                int(row) for row in self.mapping.get("manufacturing_input_rows", ())
            )),
            NumericRowSpec("front_allocation_ratios", tuple(
                int(row) for row in manufacturing.get("front_ratio_rows", {}).values()
            )),
            NumericRowSpec("production", tuple(
                int(row) for row in self.mapping.get("production", {}).values()
            )),
            NumericRowSpec("mcm", tuple(
                int(row) for row in self.mapping.get("mcm", {}).values()
            ), allow_blank=True),
            NumericRowSpec("back_raw_material", tuple(range(684, 700))),
            NumericRowSpec("sales_sources", tuple(sorted({
                int(spec[key])
                for spec in sales_groups.values()
                for key in ("quantity_row", "amount_row", "cogs_row")
                if key in spec
            }))),
            NumericRowSpec("pnl_sources", tuple(sorted(
                int(row) for row in pnl_rows.values()
            ))),
        )
        self.numeric_rows = tuple(numeric_rows or default_numeric_rows)

    def validate(
        self,
        path: str | Path,
        *,
        expected_year: int | None = None,
    ) -> PreflightReport:
        source = Path(path)
        issues: list[PreflightIssue] = []
        report = PreflightReport(path=str(source), passed=False)
        if not source.is_file():
            report.issues.append(PreflightIssue("file_missing", "업로드 워크북을 찾을 수 없습니다."))
            return report
        if source.stat().st_size > self.MAX_WORKBOOK_BYTES:
            report.issues.append(PreflightIssue(
                "file_too_large",
                "워크북이 50MB 제한을 초과했습니다.",
                expected=self.MAX_WORKBOOK_BYTES,
                observed=source.stat().st_size,
            ))
            return report
        if not zipfile.is_zipfile(source):
            report.issues.append(PreflightIssue("invalid_xlsx", "유효한 xlsx OOXML 파일이 아닙니다."))
            return report

        try:
            from openpyxl import load_workbook

            # Structural validation performs many targeted random reads.  A
            # normal worksheet is materially faster than ReadOnlyWorksheet,
            # which replays the XML stream for every random cell access.
            workbook = load_workbook(source, read_only=False, data_only=False)
        except Exception as exc:
            report.issues.append(PreflightIssue(
                "workbook_open_failed", "워크북을 열 수 없습니다.", observed=str(exc)
            ))
            return report

        try:
            data_name = next((name for name in workbook.sheetnames if name.casefold() == "data"), None)
            if data_name is None:
                issues.append(PreflightIssue(
                    "data_sheet_missing", "승인된 Data 시트를 찾을 수 없습니다.", expected="Data"
                ))
                report.issues = issues
                return report
            sheet = workbook[data_name]
            self._validate_month_columns(sheet, issues)
            found: dict[str, int] = {}
            for spec in self.anchors:
                row = self._find_anchor(sheet, spec)
                if row is None:
                    issues.append(PreflightIssue(
                        "anchor_missing",
                        f"필수 Anchor를 찾을 수 없습니다: {spec.code}",
                        expected=list(spec.labels),
                        location=f"Data row {spec.expected_row}±{spec.tolerance}",
                    ))
                    continue
                found[spec.code] = row
                if abs(row - spec.expected_row) > spec.tolerance:
                    issues.append(PreflightIssue(
                        "anchor_shifted",
                        f"Anchor 위치가 허용 범위를 벗어났습니다: {spec.code}",
                        expected=spec.expected_row,
                        observed=row,
                        location=f"Data!{row}",
                    ))
            report.anchors = found

            for block in self.blocks:
                start = found.get(block.start.code)
                stop = found.get(block.stop.code)
                if start is None or stop is None:
                    continue
                if start >= stop:
                    issues.append(PreflightIssue(
                        "anchor_order_invalid",
                        f"Anchor 블록 순서가 깨졌습니다: {block.code}",
                        expected=f"{block.start.code} < {block.stop.code}",
                        observed=f"{start} >= {stop}",
                    ))
                    continue
                outside = [row for row in block.mapped_rows if not start < row < stop]
                if outside:
                    issues.append(PreflightIssue(
                        "mapped_rows_outside_block",
                        f"매핑 행이 승인된 Anchor 블록을 벗어났습니다: {block.code}",
                        expected=f"{start + 1}..{stop - 1}",
                        observed=outside,
                    ))

            report.period_types = extract_period_types(source)
            missing_periods = [
                month for month in range(1, 13)
                if not str(report.period_types.get(str(month), "")).strip()
            ]
            if missing_periods:
                issues.append(PreflightIssue(
                    "period_type_missing",
                    "1~12월의 실적/추정/계획 구분이 모두 필요합니다.",
                    expected=list(range(1, 13)),
                    observed=missing_periods,
                    location="Data!E3:P3",
                ))
            invalid_periods = {
                month: value for month, value in report.period_types.items()
                if value and value not in _ALLOWED_PERIOD_TYPES
            }
            if invalid_periods:
                issues.append(PreflightIssue(
                    "invalid_period_type",
                    "월 구분은 실적·추정·계획만 허용됩니다.",
                    observed=invalid_periods,
                    location="Data!E3:P3",
                ))

            self._validate_numeric_sources(sheet, issues)

            report.workbook_year = infer_workbook_year(source, fallback_year=expected_year)
            if expected_year is not None and report.workbook_year != int(expected_year):
                issues.append(PreflightIssue(
                    "model_year_mismatch",
                    "워크북 연도와 등록된 모형 연도가 다릅니다.",
                    expected=int(expected_year),
                    observed=report.workbook_year,
                ))
        finally:
            workbook.close()

        report.issues = issues
        report.passed = not issues
        return report

    def require(self, path: str | Path, *, expected_year: int | None = None) -> PreflightReport:
        report = self.validate(path, expected_year=expected_year)
        if not report.passed:
            raise PreflightValidationError(report)
        return report

    @staticmethod
    def _find_anchor(sheet: Any, spec: AnchorSpec) -> int | None:
        labels = {_normalized(label) for label in spec.labels if label}
        start = max(1, spec.expected_row - spec.tolerance)
        stop = min(sheet.max_row, spec.expected_row + spec.tolerance)
        for row in range(start, stop + 1):
            for column in spec.columns:
                value = _normalized(sheet[f"{column}{row}"].value)
                if value and any(label == value or label in value for label in labels):
                    return row
        return None

    @staticmethod
    def _validate_month_columns(sheet: Any, issues: list[PreflightIssue]) -> None:
        observed: list[int | None] = []
        for column_index in range(5, 17):
            value = str(sheet.cell(2, column_index).value or "").strip()
            match = _MONTH_HEADER.search(value)
            observed.append(int(match.group(1)) if match else None)
        expected = list(range(1, 13))
        if observed != expected:
            issues.append(PreflightIssue(
                "month_header_invalid",
                "Data 시트의 E~P 열은 1월부터 12월까지 순서대로 있어야 합니다.",
                expected=expected,
                observed=observed,
                location="Data!E2:P2",
            ))

    def _validate_numeric_sources(
        self,
        sheet: Any,
        issues: list[PreflightIssue],
    ) -> None:
        for spec in self.numeric_rows:
            invalid: list[str] = []
            for row in spec.rows:
                valid_in_row = 0
                for column_index in range(5, 17):
                    cell = sheet.cell(row, column_index)
                    value = cell.value
                    if value in (None, ""):
                        continue
                    is_formula = isinstance(value, str) and value.startswith("=")
                    is_number = isinstance(value, Real) and not isinstance(value, bool)
                    if is_formula or is_number:
                        valid_in_row += 1
                    else:
                        invalid.append(cell.coordinate)
                if not spec.allow_blank and valid_in_row == 0:
                    invalid.append(f"E{row}:P{row} (all blank)")
            if invalid:
                issues.append(PreflightIssue(
                    "source_cell_not_numeric",
                    f"필수 원천셀은 숫자 또는 수식이어야 합니다: {spec.code}",
                    expected="number or formula",
                    observed=invalid[:30],
                    location=f"Data rows {','.join(str(row) for row in spec.rows)}",
                ))
