from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .workbook import GoldenWorkbook


MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MONTH_COLUMNS = {month: chr(ord("E") + month - 1) for month in range(1, 13)}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
_CHECK_FILL = PatternFill("solid", fgColor="EDEDED")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)


def _record_dict(item: Any) -> dict[str, Any]:
    if is_dataclass(item):
        return asdict(item)
    return dict(item)


def _number(value: Any) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _write_title(ws, title: str, subtitle: str | None = None) -> int:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        return 4
    return 3


def _write_headers(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"
    ws.freeze_panes = f"A{row + 1}"


def _style_data_sheet(ws, header_row: int, money_columns: Iterable[int] = (), percent_columns: Iterable[int] = ()) -> None:
    money_set, percent_set = set(money_columns), set(percent_columns)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in money_set and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
            if cell.column in percent_set and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            text = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(52, max(widths.get(cell.column, 0), len(text) + 2))
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = max(10, width)


def _formula_check(formula_cell: str, engine_cell: str, tolerance: float = 1.0) -> str:
    return f'=IF(ABS({formula_cell}-{engine_cell})<={tolerance},"PASS","CHECK")'


def _write_readme(ws, result: dict[str, Any], baseline_fx: float, comparison_fx: float) -> None:
    _write_title(ws, "손익분석 검증 엑셀", "웹 손익분석에서 사용한 입력값, 원천 셀, 계산식과 결과를 추적하기 위한 파일입니다.")
    base = result.get("baseline", {})
    comp = result.get("comparison", {})
    period = result.get("period", {})
    rows = [
        ("생성일시", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("기준 모형", f"{base.get('name', '')} / {base.get('model_type', '')} / {base.get('version', '')}"),
        ("비교 모형", f"{comp.get('name', '')} / {comp.get('model_type', '')} / {comp.get('version', '')}"),
        ("분석기간", period.get("label", "")),
        ("증감 정의", "비교 모형 - 기준 모형"),
        ("효과 부호", "손익 개선 + / 손익 악화 -"),
        ("기준 매출환율", baseline_fx),
        ("비교 매출환율", comparison_fx),
        ("영업이익 증감", result.get("operating_profit_delta", 0)),
        ("세부 효과 합계", result.get("effects_total", 0)),
        ("잔여차이", result.get("residual", 0)),
        ("정합성", "PASS" if result.get("reconciled") else "CHECK"),
    ]
    start = 4
    for index, (label, value) in enumerate(rows, start):
        ws.cell(index, 1, label).font = _BOLD
        ws.cell(index, 1).fill = _SUBHEADER_FILL
        ws.cell(index, 2, value)
    note_row = start + len(rows) + 2
    ws.cell(note_row, 1, "사용 방법").font = _BOLD
    ws.cell(note_row + 1, 1, "1. 각 검증 시트의 ‘엔진값’은 웹 화면 계산에 사용된 값입니다.")
    ws.cell(note_row + 2, 1, "2. 녹색 셀은 엑셀 수식으로 동일 계산을 재현한 값입니다.")
    ws.cell(note_row + 3, 1, "3. CHECK가 발생하면 원천 셀, 매핑 규칙, 환율 입력 또는 코드 계산을 확인합니다.")
    ws.cell(note_row + 4, 1, "4. 원천셀_추적 시트의 수식은 업로드 모형에 저장된 원본 수식이며, 값은 웹 엔진이 읽은 계산값입니다.")
    ws.cell(note_row + 5, 1, "5. MCM과 수율/사용량은 독립 손익효과로 표시하지 않습니다.")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72


def _write_pnl_and_reconciliation(ws, result: dict[str, Any]) -> None:
    header_row = _write_title(ws, "손익계산서 및 정합성", "기준·비교 손익과 손익 브리지의 정합성을 엑셀 수식으로 재확인합니다.")
    headers = ["구분", "코드", "항목", "기준 엔진값", "비교 엔진값", "증감 엔진값", "증감 엑셀수식", "검증"]
    _write_headers(ws, header_row, headers)
    row_no = header_row + 1
    for item in result.get("pnl", []):
        ws.append([
            "손익계산서", item.get("code"), item.get("item"), _number(item.get("baseline")),
            _number(item.get("comparison")), _number(item.get("delta")), None, None,
        ])
        ws.cell(row_no, 7, f"=E{row_no}-D{row_no}").fill = _FORMULA_FILL
        ws.cell(row_no, 8, _formula_check(f"G{row_no}", f"F{row_no}")).fill = _CHECK_FILL
        row_no += 1

    row_no += 1
    effect_start = row_no
    for item in result.get("effects", []):
        ws.append([
            "손익 브리지", item.get("code"), item.get("factor") or item.get("label"),
            _number(item.get("baseline")), _number(item.get("comparison")),
            _number(item.get("profit_effect")), None, None,
        ])
        if item.get("code") == "revenue":
            formula = f"=E{row_no}-D{row_no}"
        else:
            formula = f"=D{row_no}-E{row_no}"
        ws.cell(row_no, 7, formula).fill = _FORMULA_FILL
        ws.cell(row_no, 8, _formula_check(f"G{row_no}", f"F{row_no}")).fill = _CHECK_FILL
        row_no += 1
    effect_end = row_no - 1

    row_no += 1
    ws.cell(row_no, 3, "영업이익 증감 엔진값").font = _BOLD
    ws.cell(row_no, 6, _number(result.get("operating_profit_delta")))
    op_row = row_no
    row_no += 1
    ws.cell(row_no, 3, "세부 효과 합계 엔진값").font = _BOLD
    ws.cell(row_no, 6, _number(result.get("effects_total")))
    ws.cell(row_no, 7, f"=SUM(G{effect_start}:G{effect_end})").fill = _FORMULA_FILL
    ws.cell(row_no, 8, _formula_check(f"G{row_no}", f"F{row_no}")).fill = _CHECK_FILL
    total_row = row_no
    row_no += 1
    ws.cell(row_no, 3, "잔여차이 엔진값").font = _BOLD
    ws.cell(row_no, 6, _number(result.get("residual")))
    ws.cell(row_no, 7, f"=F{op_row}-G{total_row}").fill = _FORMULA_FILL
    ws.cell(row_no, 8, _formula_check(f"G{row_no}", f"F{row_no}")).fill = _CHECK_FILL
    row_no += 1
    ws.cell(row_no, 3, "최종 정합성").font = _BOLD
    ws.cell(row_no, 8, f'=IF(ABS(G{row_no-1})<=MAX(1,ABS(F{op_row})*1E-9),"PASS","CHECK")').fill = _CHECK_FILL
    _style_data_sheet(ws, header_row, money_columns=(4, 5, 6, 7))


def _write_sales(ws, sales_rows: Iterable[Any], sales_totals: dict[str, float], baseline_fx: float, comparison_fx: float) -> None:
    header_row = _write_title(ws, "판매효과 검증", "웹 화면과 동일한 수량·단가·환율 효과 산식을 셀 수식으로 재현합니다.")
    headers = [
        "제품군", "기준수량", "기준매출액", "기준단가 수식", "기준단가 엔진", "기준GP율",
        "비교수량", "비교매출액", "비교단가 수식", "비교단가 엔진", "비교GP율",
        "기준FX", "비교FX", "수량효과 수식", "수량효과 엔진", "기준외화단가 수식",
        "비교외화단가 수식", "순수단가효과 수식", "순수단가효과 엔진", "환율효과 수식",
        "환율효과 엔진", "판매효과합계 수식", "판매효과합계 엔진", "검증",
    ]
    _write_headers(ws, header_row, headers)
    row_no = header_row + 1
    first_data_row = row_no
    for source in sales_rows:
        item = _record_dict(source)
        ws.append([
            item.get("product_group", ""), _number(item.get("baseline_quantity")), _number(item.get("baseline_amount")),
            None, _number(item.get("baseline_unit_price")), _number(item.get("baseline_gross_margin_rate")),
            _number(item.get("comparison_quantity")), _number(item.get("comparison_amount")), None,
            _number(item.get("comparison_unit_price")), _number(item.get("comparison_gross_margin_rate")),
            baseline_fx, comparison_fx, None, _number(item.get("quantity_effect")), None, None, None,
            _number(item.get("pure_price_effect")), None, _number(item.get("sales_fx_effect")), None,
            _number(item.get("total_sales_effect")), None,
        ])
        formulas = {
            4: f"=IFERROR(C{row_no}/B{row_no},0)",
            9: f"=IFERROR(H{row_no}/G{row_no},0)",
            14: f"=(G{row_no}-B{row_no})*D{row_no}*F{row_no}",
            16: f"=IFERROR(D{row_no}/L{row_no},0)",
            17: f"=IFERROR(I{row_no}/M{row_no},0)",
            18: f"=G{row_no}*(Q{row_no}-P{row_no})*(L{row_no}+M{row_no})/2",
            20: f"=G{row_no}*(M{row_no}-L{row_no})*(P{row_no}+Q{row_no})/2",
            22: f"=N{row_no}+R{row_no}+T{row_no}",
            24: f'=IF(MAX(ABS(D{row_no}-E{row_no}),ABS(I{row_no}-J{row_no}),ABS(N{row_no}-O{row_no}),ABS(R{row_no}-S{row_no}),ABS(T{row_no}-U{row_no}),ABS(V{row_no}-W{row_no}))<=1,"PASS","CHECK")',
        }
        for col, formula in formulas.items():
            ws.cell(row_no, col, formula)
            ws.cell(row_no, col).fill = _FORMULA_FILL if col != 24 else _CHECK_FILL
        row_no += 1
    last_data_row = row_no - 1
    ws.cell(row_no, 1, "합계").font = _BOLD
    for col in (14, 18, 20, 22):
        ws.cell(row_no, col, f"=SUM({get_column_letter(col)}{first_data_row}:{get_column_letter(col)}{last_data_row})").fill = _FORMULA_FILL
    ws.cell(row_no, 15, _number(sales_totals.get("quantity_effect")))
    ws.cell(row_no, 19, _number(sales_totals.get("pure_price_effect")))
    ws.cell(row_no, 21, _number(sales_totals.get("sales_fx_effect")))
    ws.cell(row_no, 23, _number(sales_totals.get("total_sales_effect")))
    ws.cell(row_no, 24, f'=IF(MAX(ABS(N{row_no}-O{row_no}),ABS(R{row_no}-S{row_no}),ABS(T{row_no}-U{row_no}),ABS(V{row_no}-W{row_no}))<=1,"PASS","CHECK")').fill = _CHECK_FILL
    _style_data_sheet(ws, header_row, money_columns=(3, 4, 5, 8, 9, 10, 14, 15, 18, 19, 20, 21, 22, 23), percent_columns=(6, 11))


def _write_simple_delta_sheet(ws, title: str, subtitle: str, sections: list[tuple[str, list[dict[str, Any]]]]) -> None:
    header_row = _write_title(ws, title, subtitle)
    headers = ["구분", "코드", "항목", "기준 엔진값", "비교 엔진값", "증감 엔진값", "증감 엑셀수식", "검증", "비고"]
    _write_headers(ws, header_row, headers)
    row_no = header_row + 1
    for section, rows in sections:
        for item in rows:
            label = item.get("item") or item.get("factor") or item.get("label") or item.get("code")
            ws.append([
                section, item.get("code", ""), label, _number(item.get("baseline")),
                _number(item.get("comparison")), _number(item.get("delta")), None, None, item.get("note", ""),
            ])
            ws.cell(row_no, 7, f"=E{row_no}-D{row_no}").fill = _FORMULA_FILL
            ws.cell(row_no, 8, _formula_check(f"G{row_no}", f"F{row_no}")).fill = _CHECK_FILL
            row_no += 1
    _style_data_sheet(ws, header_row, money_columns=(4, 5, 6, 7))


def _write_material_detail(ws, result: dict[str, Any]) -> None:
    header_row = _write_title(
        ws,
        "원부재료 효과 검증",
        "엔진 Result의 제품군별 3요소를 그대로 기록합니다. 원천은 JPY 9행, 전공정 205~210행, 후공정 684~699행입니다.",
    )
    headers = [
        "제품군", "기준 원단위", "비교 원단위", "원단위 증감",
        "부직포 단가효과(환율 제외)", "부직포 엔화효과",
        "부직포 제외 원재료 효과", "원부재료 효과 합계", "3요소 엑셀합계", "검증", "계산상태",
    ]
    _write_headers(ws, header_row, headers)
    for item in (result.get("material_analysis") or {}).get("product_groups", []):
        row_no = ws.max_row + 1
        ws.append([
            item.get("product_group"), item.get("baseline_unit_cost"),
            item.get("comparison_unit_cost"), item.get("unit_cost_delta"),
            item.get("nonwoven_price_ex_fx"), item.get("nonwoven_jpy"),
            item.get("materials_ex_nonwoven"), item.get("total"), None, None,
            item.get("calculation_status"),
        ])
        ws.cell(row_no, 9, f"=SUM(E{row_no}:G{row_no})").fill = _FORMULA_FILL
        ws.cell(row_no, 10, _formula_check(f"I{row_no}", f"H{row_no}")).fill = _CHECK_FILL
    _style_data_sheet(ws, header_row, money_columns=(5, 6, 7, 8, 9))


def _write_manufacturing_detail(ws, result: dict[str, Any]) -> None:
    header_row = _write_title(
        ws,
        "생산·제조경비 효과 검증",
        "Golden Model 289~319행 계정 총액에 기준 모형 345~347행 전공정 가공비 투입비율을 적용한 엔진 Result입니다.",
    )
    headers = [
        "원천 행", "계정과목", "구분", "배부율 원천 행", "기준 전공정 배부율",
        "기준 금액", "비교 금액", "증감", "조업도 효과", "원단위 효과",
        "고정비 효과", "실현 전 효과", "재고실현율", "최종 손익효과", "발생효과 엑셀합계", "검증", "계산상태",
    ]
    _write_headers(ws, header_row, headers)
    for item in result.get("manufacturing_accounts", []):
        row_no = ws.max_row + 1
        ratios = item.get("baseline_front_ratios") or []
        ratio_display = ratios[0] if len(ratios) == 1 else ", ".join(str(value) for value in ratios)
        ws.append([
            item.get("row"), item.get("account"), item.get("classification"),
            item.get("allocation_ratio_row"), ratio_display,
            item.get("baseline_amount"), item.get("comparison_amount"), item.get("delta"),
            item.get("activity_effect"), item.get("unit_effect"), item.get("fixed_effect"),
            item.get("occurrence_effect"), item.get("inventory_realization_rate"),
            item.get("final_profit_effect"), None, None, item.get("calculation_status"),
        ])
        ws.cell(row_no, 15, f"=SUM(I{row_no}:K{row_no})").fill = _FORMULA_FILL
        ws.cell(row_no, 16, _formula_check(f"O{row_no}", f"L{row_no}")).fill = _CHECK_FILL
    _style_data_sheet(
        ws, header_row,
        money_columns=(6, 7, 8, 9, 10, 11, 12, 14, 15),
        percent_columns=(5, 13),
    )


def _write_sga_detail(ws, result: dict[str, Any]) -> None:
    header_row = _write_title(
        ws,
        "판관비 효과 검증",
        "Golden Model 계정을 개별 행으로 기록하며 고객배송 운반비와 관세의 Bridge 반영 위치를 구분합니다.",
    )
    headers = [
        "원천 행", "구역", "계정과목", "구분", "기준 금액", "비교 금액",
        "증감", "손익효과", "손익 Bridge 반영",
    ]
    _write_headers(ws, header_row, headers)
    for item in result.get("sga_accounts", []):
        ws.append([
            item.get("row"), item.get("section"), item.get("account"),
            item.get("classification"), item.get("baseline_amount"),
            item.get("comparison_amount"), item.get("delta"),
            item.get("profit_effect"), item.get("bridge_position"),
        ])
    _style_data_sheet(ws, header_row, money_columns=(5, 6, 7, 8))


def _write_formula_catalog(ws) -> None:
    header_row = _write_title(ws, "수식 정의", "분석 엔진과 검증 엑셀에서 사용하는 주요 산식 및 처리 원칙입니다.")
    headers = ["영역", "항목", "산식", "부호·처리 기준", "코드 위치"]
    _write_headers(ws, header_row, headers)
    rows = [
        ("공통", "증감", "비교값 - 기준값", "단순 증감", "forecast/comparison.py"),
        ("판매", "수량효과", "(비교수량-기준수량)×기준단가×기준GP율", "개선 + / 악화 -", "forecast/sales_comparison.py"),
        ("판매", "순수 단가효과", "비교수량×(비교외화단가-기준외화단가)×(기준FX+비교FX)÷2", "환율효과와 합계가 원화 단가효과에 일치", "forecast/sales_comparison.py"),
        ("판매", "매출환율효과", "비교수량×(비교FX-기준FX)×(기준외화단가+비교외화단가)÷2", "KRW/USD", "forecast/sales_comparison.py"),
        ("원부재료", "분해 원칙", "부직포 단가(환율 제외)+부직포 엔화+부직포 제외 원재료", "MCM·수율/사용량 독립효과 금지", "forecast/analysis/material_effects.py"),
        ("생산", "조업도 기준", "SAP 수불부 생산입고", "MES는 정합성 확인 보조", "분석 설정"),
        ("제조경비", "외주가공비 수량", "일반 외주가공 대상 수량", "MCM 관련 수량 제외", "분석 설정"),
        ("손익 브리지", "비용 효과", "기준 비용-비교 비용", "비용 감소는 손익 개선 +", "forecast/comparison.py"),
        ("정합성", "잔여차이", "영업이익 증감-세부효과 합계", "허용오차 이내 PASS", "forecast/comparison.py"),
    ]
    for row in rows:
        ws.append(row)
    _style_data_sheet(ws, header_row)


def _mapping_specs(mapping: dict[str, Any]) -> list[tuple[str, str, str, Any]]:
    output: list[tuple[str, str, str, Any]] = []
    for code, row in mapping.get("pnl_rows", {}).items():
        output.append(("손익계산서", code, mapping.get("pnl_labels", {}).get(code, code), row))
    for code, spec in mapping.get("products", {}).items():
        output.append(("판매 제품", f"{code}.quantity", f"{spec.get('label', code)} 수량", spec.get("quantity_row")))
        output.append(("판매 제품", f"{code}.amount", f"{spec.get('label', code)} 매출액", spec.get("amount_row")))
    for code, spec in mapping.get("sales_groups", {}).items():
        output.append(("판매 제품군", f"{code}.quantity", f"{spec.get('label', code)} 수량", spec.get("quantity_row")))
        output.append(("판매 제품군", f"{code}.amount", f"{spec.get('label', code)} 매출액", spec.get("amount_row")))
        output.append(("판매 제품군", f"{code}.cogs", f"{spec.get('label', code)} 매출원가", spec.get("cogs_row")))
    for code, row in mapping.get("production_rows", {}).items():
        output.append(("생산", code, mapping.get("production_labels", {}).get(code, code), row))
    for code, row in mapping.get("mcm_rows", {}).items():
        output.append(("MCM", code, mapping.get("mcm_labels", {}).get(code, code), row))
    for code, spec in mapping.get("cost_rows", {}).items():
        output.append(("비용 요약", code, mapping.get("cost_labels", {}).get(code, code), spec))
    for code, spec in mapping.get("effect_rows", {}).items():
        output.append(("손익 브리지", code, mapping.get("effect_labels", {}).get(code, code), spec))
    return output


def _analysis_source_specs(payload: dict[str, Any]) -> list[tuple[str, str, str, Any]]:
    adapter = payload.get("analysis_adapter", {})
    material = adapter.get("material", {})
    front = material.get("front_process", {})
    back = material.get("back_process", {})
    output: list[tuple[str, str, str, Any]] = []
    if material.get("jpy_fx_row"):
        output.append(("원부재료", "jpy_fx", "엔화환율(KRW/JPY)", material["jpy_fx_row"]))
    for key, label in (
        ("nonwoven_quantity_row", "전공정 부직포 생산출고 수량"),
        ("nonwoven_amount_row", "전공정 부직포 생산출고 금액"),
        ("nonwoven_unit_row", "전공정 부직포 생산출고 단가"),
        ("other_quantity_row", "부직포 제외 전공정 원재료 생산출고 수량"),
        ("other_amount_row", "부직포 제외 전공정 원재료 생산출고 금액"),
        ("other_unit_row", "부직포 제외 전공정 원재료 생산출고 단가"),
        ("total_amount_row", "전공정 원재료 생산출고 합계"),
    ):
        if front.get(key):
            output.append(("원부재료", key, label, front[key]))
    if back.get("source_start_row") and back.get("source_end_row"):
        for row in range(int(back["source_start_row"]), int(back["source_end_row"]) + 1):
            output.append(("원부재료", f"back_process_{row}", f"후공정 원재료 생산출고 {row}행", row))
    account_discovery = adapter.get("account_discovery", {})
    if account_discovery.get("manufacturing_start_marker"):
        for row in range(289, 320):
            output.append(("제조경비", f"manufacturing_{row}", f"제조경비 명세 {row}행", row))
    for key, row in adapter.get("manufacturing", {}).get("front_ratio_rows", {}).items():
        output.append(("제조경비", f"front_ratio_{key}", f"전공정 가공비 투입비율({key})", row))
    return output


def _expand_spec(spec: Any) -> list[tuple[int, int, str]]:
    if isinstance(spec, int):
        return [(spec, 1, "direct")]
    if isinstance(spec, list):
        return [(int(row), 1, "add") for row in spec]
    if isinstance(spec, dict):
        rows = [(int(row), 1, "add") for row in spec.get("add", [])]
        rows.extend((int(row), -1, "subtract") for row in spec.get("subtract", []))
        return rows
    return []


def _write_source_trace(
    ws,
    baseline_path: str | Path,
    comparison_path: str | Path,
    mapping_path: str | Path,
    months: tuple[int, ...],
) -> None:
    header_row = _write_title(ws, "원천셀 추적", "모형별·월별로 실제 읽은 셀, 저장 수식, 계산값과 집계 부호를 표시합니다.")
    headers = ["모형", "영역", "지표코드", "지표명", "월", "시트", "셀", "집계부호", "집계규칙", "원본수식", "엔진 사용값"]
    _write_headers(ws, header_row, headers)
    payload = json.loads(Path(mapping_path).read_text(encoding="utf-8"))
    mapping = payload["comparison"]
    specs = [*_mapping_specs(mapping), *_analysis_source_specs(payload)]
    models = [("기준", GoldenWorkbook(baseline_path)), ("비교", GoldenWorkbook(comparison_path))]
    for side, workbook in models:
        for domain, code, label, spec in specs:
            for row_number, sign, rule in _expand_spec(spec):
                for month in months:
                    cell_ref = f"{MONTH_COLUMNS[int(month)]}{row_number}"
                    ws.append([
                        side, domain, code, label, f"{month}월", "Data", cell_ref, sign, rule,
                        workbook.formulas.get(cell_ref, ""), _number(workbook.value(cell_ref)),
                    ])
    _style_data_sheet(ws, header_row, money_columns=(11,))


def build_comparison_audit_workbook(
    *,
    result: dict[str, Any],
    sales_rows: Iterable[Any],
    sales_totals: dict[str, float],
    baseline_fx: float,
    comparison_fx: float,
    baseline_path: str | Path,
    comparison_path: str | Path,
    mapping_path: str | Path,
) -> bytes:
    """Build a formula-bearing audit workbook for the current comparison result."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    _write_readme(workbook.create_sheet("README"), result, baseline_fx, comparison_fx)
    _write_pnl_and_reconciliation(workbook.create_sheet("손익_정합성"), result)
    _write_sales(workbook.create_sheet("판매효과_검증"), sales_rows, sales_totals, baseline_fx, comparison_fx)

    _write_material_detail(workbook.create_sheet("원부재료_검증"), result)
    _write_manufacturing_detail(workbook.create_sheet("생산제조경비_검증"), result)
    _write_sga_detail(workbook.create_sheet("판관비_검증"), result)
    _write_formula_catalog(workbook.create_sheet("수식_정의"))
    months = tuple(int(month) for month in result.get("period", {}).get("months", ()))
    _write_source_trace(
        workbook.create_sheet("원천셀_추적"), baseline_path, comparison_path, mapping_path, months,
    )

    output = BytesIO()
    workbook.save(output)
    payload = output.getvalue()

    probe = load_workbook(BytesIO(payload), data_only=False, read_only=True)
    required = {
        "README", "손익_정합성", "판매효과_검증", "원부재료_검증",
        "생산제조경비_검증", "판관비_검증", "수식_정의", "원천셀_추적",
    }
    missing = required.difference(probe.sheetnames)
    if missing:
        raise ValueError(f"검증 엑셀 필수 시트 누락: {sorted(missing)}")
    if not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for row in probe["판매효과_검증"].iter_rows()
        for cell in row
    ):
        raise ValueError("판매효과 검증 수식이 생성되지 않았습니다.")
    return payload
